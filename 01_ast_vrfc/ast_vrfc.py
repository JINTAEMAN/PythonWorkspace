# -*- coding: utf-8 -*-	    # 문자 인코딩(한글 사용) 
# ! /ast_vrfc.py		    # 자산 검증 파일 

from openpyxl import load_workbook # 파일 불러오기 --> [기초 Data]
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles import PatternFill
import os       # 파이씬 os 제어 Lib
import sys	    # 파이씬 인터프리터 제어 Lib
import pyautogui    # 마우스와 키보드 제어 Lib
import time     # 시간 Lib
from datetime import datetime   # datetime Lib 
import calendar  #   calendar Lib
import requests
from bs4 import BeautifulSoup

def getMonthRage(year, month):      # 월 날짜 범위 설정 함수()
    date = datetime(year=year, month=month, day=1).date()
    monthrange = calendar.monthrange(date.year, date.month)
    first_day = calendar.monthrange(date.year, date.month)[0]
    last_day = calendar.monthrange(date.year, date.month)[1] 
    # print("[@_T] ■■■ [/ast_vrfc.py] [getMonthRage]==> [T_91] [monthrange]"+ str(monthrange) )

    return monthrange

def readParameters():    # 파일에서 파라미터 가져오는 함수()
    print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_01]")

    file = open("01_pram_properties.txt", 'rt', encoding='utf-8-sig')	# properties.txt 파일 내용 ---> 2023.09[자산 년월(오늘 년월)]
    # print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_51] [file]"+ str(file) )

    parameters = file.read().split(";")      # 자산 년월 parameters
    print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_91] [parameters]"+ str(parameters) )
    print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_92] [자산 년월]"+ str(parameters[0]) )

    return parameters
# # ---------------------------------------------------------------------------------------------------------------------->

now_ym = str(datetime.today().date()).replace('-','.')     # 오늘 년월(년.월) (2023.08.31)
now_ym = now_ym[0:7]     # 자산 년월(2023.08) 
# ---------------------------------------------------------------------------------------------------------------------->

parameters = readParameters()   # 파일에서 파라미터 가져오는 함수()
astYYM = str(parameters[0])     # 자산 년월
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_2] [parameters[0]]"+ str(parameters[0]) +"[자산 년월]"+ str(astYYM) ) 

if str(astYYM) == None :   # 자산 년월 미입력 이면
    result = pyautogui.alert("자산 년월을 입력하세요. 예) 2023.08", title='[자산 년월 입력 오류]', button='OK')
    sys.exit()    # 종료

astYYM = astYYM         # 자산 년월 ■■■■■■ (2023.08)
astYM = astYYM[2:8]     # 자산 년월(23.08)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_3] [자산 년월_7]"+ str(astYYM) +"[자산 년월_5]"+ str(astYM) )

# urlPath = "01_ast_vrfc/"   # URL 경로 
urlPath = ""   # URL 경로
openFileNm = "01. 자산 검증("+ astYM +").xlsx"   # 오픈 파일 명(01. 자산 검증(23.08).xlsx) 
wb = load_workbook(urlPath + openFileNm, data_only=True)    # 오픈 파일을 wb을 불러옴(data_only=True: 수식이 아닌 실제 데이터를 가지고 옴)

today = datetime.today().date()
month_range = getMonthRage(today.year, today.month)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_4] [today]"+ str(today) +"[month_range]"+ str(month_range) )

lastLastDt = str(today.replace(day = month_range[1])).replace('-','.')       # 자산 마지막 일 ■■■■■■ "2023.08.31"
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_5] [자산 마지막 일]"+ str(lastLastDt) )

# 00. 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 00. 음영 색 지정
orangeFill = PatternFill(start_color='F79646', end_color='F79646', fill_type='solid')   # 오렌지 색
orangeWeekFill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')    # 연한 오렌지
orangeWeek2Fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')    # 연한 오렌지2
orangeWeek3Fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')    # 연한 오렌지3 
grayFill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')     # 회 색
grayDarkFill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')    # 진한 회색 
grayDark2Fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')    # 진한 회색2  
blueFill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')     # 하늘 색
blueDarkFill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')     # 진한 하늘 색
blueDark2Fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')     # 진한 하늘 색2
blueDark3Fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')    # 진한 파란 색
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # 노란 색
violetFill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')   # 보라 색
greenWeekFill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')   # 연한 녹색
greenFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')    # 녹색
# ---------------------------------------------------------------------------------------------------------------------->
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


ws_pay = wb["01. 급여"]   # "01. 급여" Sheet에 접근 @@@@@@@@ ===========>

# 종목 코드 리스트
codes = [
    '005380'   # 현대차     
    , '068270' # 셀트리온
    , '096770'  # SK이노베이션
]

row = 0  

for code in codes:   # "01. 급여" Sheet에서 월별 보유 주식 현황 Data 줄
    if code == "068270":     # 증권 종목이 '셀트리온' 이면  
        posses_qty2 = ws_pay[f'X{row+3}'].value  	    # 주식(셀트리온) 보유 수량
        ave_prchs_amt2 =  ws_pay[f'Z{row+3}'].value     # 주식(셀트리온) 평균 매입가
        prchs_amt2 =  ws_pay[f'AA{row+3}'].value        # 주식(셀트리온) 매입 금액 		
    elif code == "096770":   # 증권 종목이 'SK이노베이션' 이면  
        posses_qty3 = ws_pay[f'X{row+3}'].value  	    # 주식(SK이노베이션) 보유 수량
        ave_prchs_amt3 =  ws_pay[f'Z{row+3}'].value     # 주식(SK이노베이션) 평균 매입가
        prchs_amt3 =  ws_pay[f'AA{row+3}'].value        # 주식(SK이노베이션) 매입 금액 
    else :      		
        posses_qty1 = ws_pay[f'X{row+3}'].value  	    # 주식(현대차) 보유 수량
        ave_prchs_amt1 =  ws_pay[f'Z{row+3}'].value     # 주식(현대차) 평균 매입가
        prchs_amt1 =  ws_pay[f'AA{row+3}'].value        # 주식(현대차) 매입 금액

    if row == 0:
        shinhanBValue = ws_pay[f'AG{row+3}'].value      # 은행 결산@@ 신한은행 금액
    elif row == 1:
        kakaoBValue = ws_pay[f'AG{row+3}'].value  	    # 은행 결산@@ 카카오 뱅크 금액
    else :      		
        shinhybBValue = ws_pay[f'AG{row+3}'].value     # 은행 결산@@ 신협 예금 금액

    row = row + 1   

add_row = 0           # 줄 추가(add_row): "01. 급여" Sheet에서 월별 보유 주식 현황 Data가 추가될 경우 대비
row = add_row + 6     # "01. 급여" Sheet에서 합계 줄
posses_qty_sum = ws_pay[f'X{row}'].value  	    # 보유 수량 합계
ave_prchs_amt_sum = ws_pay[f'Z{row}'].value  	# 평균 매입가 합계
stockPrchsAmt_sum = ws_pay[f'AA{row}'].value  	# 매입금액 합계(주식 총 매입 금액)
sokSavBValue = ws_pay[f'AG{row}'].value  	    # 은행 결산@@ OK 저축 은행 예금 금액

row = add_row + 7     # "01. 급여" Sheet에서 비 보험 줄
insuSonValue =  ws_pay[f'X{row}'].value     # 21. 실비 보험(현대해상, 진수종)
insuValue = ws_pay[f'AA{row}'].value        # 22. 실비 보험(한화손해보험, 잔태만)
realtyValue = ws_pay[f'AD{row}'].value      # 3. 부동산(현아트빌 404호) 금액 

row = add_row + 8     # "01. 급여" Sheet에서 퇴직 연금 줄
pensionIRP = ws_pay[f'AD{row}'].value       # 91. 퇴직 연금(개인형 IPR) 금액
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_03_1] [보유 수량 합계]"+ str(posses_qty_sum) +"[신한은행 금액)]"+ str(shinhanBValue) +"[퇴직 연금(IPR)]"+ str(pensionIRP))
# ---------------------------------------------------------------------------------------------------------------------->
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


ws_stock = wb["91. 주식"]   # "91. 주식" Sheet에 접근 @@@@@@@@ ===========> 
prev_B3 = ws_stock["B3"].value   # 이전 일자(예) 2023.08.31)
prev_B3 = prev_B3[0:7]     # 이전 일자[년월) (2023.08)

ws_stock.insert_rows(3, 8)       # 월별 자산 제목(3번째 줄 위치에 8줄을 추가)
ws_stock.merge_cells("A8:B8")    # A8부터 B9까지 셀을 싱글 셀로 병합
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_2] [91. 주식 Sheet] [오늘 년월]"+ now_ym.lstrip() +"[오늘 년월_타입]"+ str(type(now_ym)) +"[오늘 년월_갯수]"+ str(len(now_ym.lstrip())) )
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_3] [91. 주식 Sheet] [자산 년월]"+ astYYM.lstrip() +"[자산 년월_타입]"+ str(type(astYYM)) +"[자산 년월_갯수]"+ str(len(astYYM.lstrip())) )

if str(astYYM.strip()) != str(now_ym.strip()) :   # 입력한 자산 년월와 오늘 년월이 다르면
    result = pyautogui.alert("[91. 주식 Sheet] 입력한 자산 년월["+ str(astYYM) +"]와 오늘 년월["+ str(now_ym) +"]이 다릅니다.", title='[자산 년월 입력 오류]', button='OK')
    sys.exit()    # 종료
else :   # 입력한 자산 년월와 오늘 년월이 같으면
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_4_2] [자산 년월]"+ str(astYYM) +"[오늘 년월]"+ str(now_ym) )

    if str(astYYM) == str(prev_B3) :   # 입력한 자산 년월와 이전 자산 년월이 같으면
        result = pyautogui.alert("[91. 주식 Sheet] 입력한 자산 년월["+ str(astYYM) +"]이 이미 존재합니다.", title='[자산 년월 입력 오류]', button='OK')
        sys.exit()    # 종료
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_3] [입력한 자산 년월]"+ str(astYYM) +"[오늘 년월]"+ str(now_ym) +"[이전 자산 년월]"+ str(prev_B3) )
# ---------------------------------------------------------------------------------------------------------------------->

print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_5]" ) 

prchs_amt = 0               # 매입 금액
prchs_amt_sum = 0           # 매입 금액 합계
evl_amt = 0                 # 평가 금액
evl_amt_sum = 0             # 평가 금액 합계
evl_profitLoss_sum = 0      # 평가 손익 합계 
evl_profitLoss = 0          # 평가 손익
profitLossRate_sum = 0      # 손익률 합계
row = 0  

for code in codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"    # 네이버 증권 종목(f: 포맷 문자열, 리터럴 사용)
    # url = "https://finance.naver.com/item/sise.naver?code="+ code    # 네이버 증권 종목     
    # url = 'https://finance.naver.com/item/sise.naver?code=005380'    # 네이버 증권 종목 ===> TEST @@@ ===>

    response = requests.get(url)     # 웹 사이트 열기

    if response.status_code == 200:        
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        price = soup.select_one("#_nowVal").text    # 현재가 ==>sCSS_SELECTOR
        # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_4] [row_번째]"+ str(row) +"[현재가]"+ str(price) +"[url_주소]"+ str(url) )  

        price = price.replace(',', '')    
        # price =  5000000    # 현재가  TEST @@@ ===>
        # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_5] [row_번째] [soup]"+ str(soup) )

    else : 
        print(response.status_code) 
    
    if code == "068270":     # 증권 종목이 '셀트리온' 이면 
        stock_nm = "셀트리온"         # 주식명
        posses_qty = posses_qty2    # 보유수량
        prchs_amt = prchs_amt2      # 매입 금액
        ave_prchs_amt = ave_prchs_amt2   # 평균 매입가
    elif code == "096770":   # 증권 종목이 'SK이노베이션' 이면 
        stock_nm = "SK이노베이션"    # 주식명
        posses_qty = posses_qty3    # 보유수량  
        prchs_amt = prchs_amt3      # 매입 금액
        ave_prchs_amt = ave_prchs_amt3   # 평균 매입가
    else :
        stock_nm = "현대차"          # 주식명
        posses_qty = posses_qty1    # 보유수량
        prchs_amt = prchs_amt1      # 매입 금액
        ave_prchs_amt = ave_prchs_amt1   # 평균 매입가
    
    evl_amt = int(posses_qty) * int(price)     # 평가금액 = 보유수량 * 현재가(=P5*Q5) 
    evl_profitLoss = (evl_amt) - prchs_amt       # 평가손익 = 평가금액 - 매입금액(=T14-S14)
    print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_6] [row_번째]"+ str(row) +"[매입 금액]"+ str(prchs_amt) +"[보유수량]"+ str(posses_qty) +"[현재가]"+ str(price) ) 

    ws_stock[f'A{row+5}'] = code                # 주식코드
    ws_stock[f'A{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 주식코드(02. 스타일 적용)     
    ws_stock[f'A{row+5}'].border = thin_border   # 주식코드(04. 테두리 적용)
    ws_stock[f'B{row+5}'] = stock_nm            # 주식명
    ws_stock[f'B{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 주식명(02. 스타일 적용)
    ws_stock[f'B{row+5}'].border = thin_border   # 주식코드(04. 테두리 적용)
    ws_stock[f'C{row+5}'] = int(posses_qty)     # 매입수량
    ws_stock[f'C{row+5}'].font = Font(name="돋움", bold=False, size=9)   # bold(02. 스타일 적용)
    ws_stock[f'C{row+5}'].border = thin_border   # 주식코드(04. 테두리 적용)
    ws_stock[f'D{row+5}'] = format(int(price), ',')  # 현재가 
    ws_stock[f'D{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 현재가(02. 스타일 적용)
    ws_stock[f'D{row+5}'].alignment = Alignment(horizontal='right', vertical='center')     # 현재가(03. alignment 적용)
    ws_stock[f'D{row+5}'].border = thin_border   # 현재가(04. 테두리 적용)
    ws_stock[f'D{row+5}'].fill = yellowFill   # 현재가(05. 배경색 설정)
    ws_stock[f'E{row+5}'] = format(int(ave_prchs_amt), ',')    # 평균 매입가
    ws_stock[f'E{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 평균 매입가(02. 스타일 적용)
    ws_stock[f'E{row+5}'].alignment = Alignment(horizontal='right', vertical='center')     # 평균 매입가(03. alignment 적용)
    ws_stock[f'E{row+5}'].border = thin_border   # 평균 매입가(04. 테두리 적용)
    ws_stock[f'F{row+5}'] = format(int(prchs_amt), ',')        # 매입 금액 
    ws_stock[f'F{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 매입 금액(02. 스타일 적용)
    ws_stock[f'F{row+5}'].alignment = Alignment(horizontal='right', vertical='center')     # 매입 금액(03. alignment 적용)
    ws_stock[f'F{row+5}'].border = thin_border   # 매입 금액(04. 테두리 적용) 
    ws_stock[f'G{row+5}'] = format(int(evl_amt), ',')          # 평가금액 = 보유수량 * 현재가(=P5*Q5) 
    ws_stock[f'G{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 평가금액(02. 스타일 적용)
    ws_stock[f'G{row+5}'].alignment = Alignment(horizontal='right', vertical='center')     # 평가금액(03. alignment 적용)
    ws_stock[f'G{row+5}'].border = thin_border   # 평가금액(04. 테두리 적용)
    ws_stock[f'G{row+5}'].fill = orangeWeek3Fill   # 평가금액(05. 배경색 설정)
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_3] [row_번째]"+ str(row) +"[종목 코드]"+ code +"[01. 보유수량]"+ str(int(ws_stock[f'P{row+5}'].value)) +"[02. 현재가]"+ str(price) +"[03. 평가금액]"+ str(evl_amt) ) 

    ws_stock[f'H{row+5}'] = format(int(evl_profitLoss), ',')    # 평가손익 
    ws_stock[f'H{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 평가손익(02. 스타일 적용)
    ws_stock[f'H{row+5}'].alignment = Alignment(horizontal='right', vertical='center')     # 평가손익(03. alignment 적용)
    ws_stock[f'H{row+5}'].border = thin_border   # 평가손익(04. 테두리 적용) 
    ws_stock[f'H{row+5}'].fill = orangeWeek3Fill   # 평가손익(05. 배경색 설정)

    profitLossRate = ((evl_amt / prchs_amt) - 1) * 100       # 손익률 = ((평가금액 / 매입금액) -1) *100 (=(T14/S14)-1) * 100) 
    ws_stock[f'I{row+5}'] = f'{profitLossRate:.2f}'         # 손익률 = ((평가금액 / 매입금액) -1) *100 (=(T14/S14)-1) * 100)
    ws_stock[f'I{row+5}'].font = Font(name="돋움", bold=False, size=9)   # 손익률(02. 스타일 적용)
    ws_stock[f'I{row+5}'].alignment = Alignment(horizontal='right', vertical='center')     # 손익률(03. alignment 적용)
    ws_stock[f'I{row+5}'].border = thin_border   # 손익률(04. 테두리 적용)
    ws_stock[f'I{row+5}'].fill = orangeWeek3Fill   # 손익률(05. 배경색 설정)
    
    ws_stock[f'J{row+5}'].border = thin_border   # J5 필드 border 설징(빈칸) 
    ws_stock[f'K{row+5}'].border = thin_border   # K5 필드 border 설징(빈칸) 
    ws_stock[f'L{row+5}'].border = thin_border   # L5 필드 border 설징(빈칸) 
    ws_stock[f'M{row+5}'].border = thin_border   # M5 필드 border 설징(빈칸)
    
    evl_amt_sum = evl_amt_sum + evl_amt     # 평가 금액 합계
    evl_profitLoss_sum = evl_profitLoss_sum + evl_profitLoss      # 평가손익 합계
    prchs_amt_sum = prchs_amt_sum + prchs_amt      # 매입 금액 합계 
    print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_7] [row_번째]"+ str(row) +"[종목 코드]"+ code +"[00. 손익률]"+ str(profitLossRate) +"[01. 보유수량]"+ str(posses_qty) +"[02. 매입 금액]"+ str(prchs_amt) +"[02. 현재가]"+ str(price) +"[03. 평가금액]"+ str(evl_amt) +"[04. 평가손익]"+ str(evl_profitLoss) ) 
    
    row = row + 1
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_8]" )

profitLossRate_sum = ((evl_amt_sum / prchs_amt_sum) - 1) * 100     # 손익률 합계 = ((평가금액 합계 / 매입금액 합계) -1) *1 00 
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_9] [01. 평가금액 합계]"+ str(evl_amt_sum) +"[02. 매입금액 합계]"+ str(prchs_amt_sum) +"[03. 손익률 합계]"+ str(profitLossRate_sum) +"[05. 평가손익 합계]"+ str(evl_profitLoss) ) 

# 01. 값 쓰기 
ws_stock["A3"].value = "일자"                       # A3 필드에 값 쓰기(일자)  -------->
ws_stock["B3"].value = lastLastDt                  # B3 필드에 값 쓰기(일자 Value) 
ws_stock["C3"].value = "주식 평가 금액"             # C3 필드에 값 쓰기(주식 평가 금액)
ws_stock["D3"] = format(evl_amt_sum, ',')          # 평가금액 합계
ws_stock["E3"].value = "총 매입 금액"                # E3 필드에 값 쓰기 
ws_stock["F3"] = format(prchs_amt_sum, ',')         # 매입 단가 합계
ws_stock["K3"] = format(evl_profitLoss_sum, ',')    # 평가손익 합계
ws_stock["H3"].value = "보유예탁자산 증감"            # H3 필드에 값 쓰기 
ws_stock["J3"].value = "평가 손익 합계"               # J3 필드에 값 쓰기 
ws_stock["K3"] = format(evl_profitLoss_sum, ',')     # K3 필드에 값 쓰기(평가 손익 합계 Value)
ws_stock["L3"].value = "총평가 손익률(%)"             # L3 필드에 값 쓰기 
ws_stock["M3"] = f'{profitLossRate_sum:.2f}'    # M3 필드에 값 쓰기(총평가 손익률(%) Value)
ws_stock["A4"].value = "주식코드"               # A4 필드에 값 쓰기(주식코드) Head  -------->
ws_stock["B4"].value = "주식명"                 # B4 필드에 값 쓰기 
ws_stock["C4"].value = "매입수량"               # C4 필드에 값 쓰기 
ws_stock["D4"].value = "현재가"                 # D4 필드에 값 쓰기 
ws_stock["E4"].value = "매입금액"               # E4 필드에 값 쓰기 
ws_stock["F4"].value = "현재가"                 # F4 필드에 값 쓰기 
ws_stock["G4"].value = "평가금액"               # G4 필드에 값 쓰기 
ws_stock["H4"].value = "평가손익"               # H4 필드에 값 쓰기 
ws_stock["I4"].value = "손익률(%)"              # I4 필드에 값 쓰기 
ws_stock["J4"].value = "비고"                   # J4 필드에 값 쓰기 
ws_stock["L4"].value = "목표가"                 # L4 필드에 값 쓰기 
ws_stock["M4"].value = "손절가"                 # M4 필드에 값 쓰기

row = 0
add_row = 0           # 줄 추가(add_row): "91. 주식" Sheet에서 보유 주식 현황 Data가 추가될 경우 대비
row = add_row + 8     # "91. 주식" Sheet에서 합계 줄
ws_stock[f'A{row}'].value = "합계"                   # A8 필드에 값 쓰기(합계)  -------->
ws_stock[f'C{row}'].value = posses_qty_sum 		# 매입수량 합계 
ws_stock[f'E{row}'].value = format(ave_prchs_amt_sum, ',') 		# 평균 매입가(매입단가) 합계
ws_stock[f'F{row}'].value = format(prchs_amt_sum, ',')  		# 매입 단가 합계 
ws_stock[f'G{row}'].value = format(evl_amt_sum, ',') 		    # 평가금액 합계 
ws_stock[f'H{row}'].value = format(evl_profitLoss_sum, ',')     # 평가손익 합계  ===>  
ws_stock[f'I{row}'].value = f'{profitLossRate_sum:.2f}' 		# 손익률(%) 합계

# 02. 스타일 적용드 스타일 적용(글자 색은 빨갛게, 이탤릭, 두껍게 적용)
ws_stock["A3"].font = Font(name="돋움", bold=False, size=9)   # A3 필드(자산 년월)  -------->
ws_stock["B3"].font = Font(name="돋움", bold=False, size=9)    # B3 필드 
ws_stock["C3"].font = Font(name="돋움", bold=False, size=9)    # C3 필드 
ws_stock["D3"].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # D3 필드(주식 평가 금액 Value)
ws_stock["E3"].font = Font(name="돋움", bold=False, size=9)    # E3 필드
ws_stock["F3"].font = Font(name="돋움", bold=True, size=9, color="FF0000")     # F3 필드(총 매입 금액 Value)
ws_stock["H3"].font = Font(name="돋움", bold=False, size=9)    # H3 필드
ws_stock["J3"].font = Font(name="돋움", bold=False, size=9)    # J3 필드(평가 손익 합계)
ws_stock["K3"].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # K3 필드(평가 손익 합계 Value)
ws_stock["L3"].font = Font(name="돋움", bold=False, size=9)    # L3 필드(총평가 손익률(%))
ws_stock["M3"].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # M3 필드(총평가 손익률(%) Value)
ws_stock["A4"].font = Font(name="돋움", bold=False, size=9)   # A4 필드(주식코드) Head  -------->
ws_stock["B4"].font = Font(name="돋움", bold=False, size=9)   # B4 필드
ws_stock["C4"].font = Font(name="돋움", bold=False, size=9)   # C4 필드
ws_stock["D4"].font = Font(name="돋움", bold=False, size=9)   # D4 필드
ws_stock["E4"].font = Font(name="돋움", bold=False, size=9)   # E4 필드
ws_stock["F4"].font = Font(name="돋움", bold=False, size=9)   # F4 필드
ws_stock["G4"].font = Font(name="돋움", bold=False, size=9)   # G4 필드
ws_stock["H4"].font = Font(name="돋움", bold=False, size=9)   # H4 필드
ws_stock["I4"].font = Font(name="돋움", bold=False, size=9)   # I4 필드
ws_stock["J4"].font = Font(name="돋움", bold=False, size=9)   # J4 필드
ws_stock["L4"].font = Font(name="돋움", bold=False, size=9)   # M4 필드
ws_stock["M4"].font = Font(name="돋움", bold=False, size=9)   # AB 필드(손절가)
ws_stock[f'A{row}'].font = Font(name="돋움", bold=False, size=9)  		# A8 필드(합계)  -------->
ws_stock[f'C{row}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # C8 필드
ws_stock[f'E{row}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # E8 필드
ws_stock[f'F{row}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # F8 필드
ws_stock[f'G{row}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")     # G8 필드
ws_stock[f'H{row}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")     # H8 필드
ws_stock[f'I{row}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")     # I8 필드

# 03. alignment 적용
ws_stock["A3"].alignment = Alignment(horizontal='center', vertical='center')    # A3 필드 alignment 설정 -------->
ws_stock["B3"].alignment = Alignment(horizontal='center', vertical='center')    # B3 필드
ws_stock["C3"].alignment = Alignment(horizontal='center', vertical='center')    # C3 필드
ws_stock["D3"].alignment = Alignment(horizontal='right', vertical='center')     # D3 필드
ws_stock["E3"].alignment = Alignment(horizontal='center', vertical='center')     # E3 필드
ws_stock["F3"].alignment = Alignment(horizontal='right', vertical='center')      # F3 필드
ws_stock["H3"].alignment = Alignment(horizontal='center', vertical='center')     # H3 필드
ws_stock["J3"].alignment = Alignment(horizontal='center', vertical='center')     # J3 필드(평가 손익 합계)
ws_stock["K3"].alignment = Alignment(horizontal='right', vertical='center')      # K3 필드(평가 손익 합계 Value)
ws_stock["L3"].alignment = Alignment(horizontal='center', vertical='center')     # L3 필드(총평가 손익률(%))
ws_stock["M3"].alignment = Alignment(horizontal='right', vertical='center')      # M3 필드(총평가 손익률(%) Value)
ws_stock["A4"].alignment = Alignment(horizontal='center', vertical='center')     # A4 필드(주식코드) Head  -------->
ws_stock["B4"].alignment = Alignment(horizontal='center', vertical='center')     # B4 필드
ws_stock["C4"].alignment = Alignment(horizontal='center', vertical='center')     # C4 필드
ws_stock["D4"].alignment = Alignment(horizontal='center', vertical='center')     # D4 필드
ws_stock["E4"].alignment = Alignment(horizontal='center', vertical='center')     # E4 필드
ws_stock["F4"].alignment = Alignment(horizontal='center', vertical='center')     # F4 필드
ws_stock["G4"].alignment = Alignment(horizontal='center', vertical='center')     # G4 필드
ws_stock["H4"].alignment = Alignment(horizontal='center', vertical='center')     # H4 필드
ws_stock["I4"].alignment = Alignment(horizontal='center', vertical='center')     # I4 필드
ws_stock["J4"].alignment = Alignment(horizontal='center', vertical='center')     # A4 필드
ws_stock["L4"].alignment = Alignment(horizontal='center', vertical='center')     # L4 필드
ws_stock["M4"].alignment = Alignment(horizontal='center', vertical='center')     # M4 필드(손절가)
ws_stock[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')    # A8 필드(합계)  -------->
ws_stock[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')     # C8 필드
ws_stock[f'E{row}'].alignment = Alignment(horizontal='right', vertical='center')     # E8 필드
ws_stock[f'F{row}'].alignment = Alignment(horizontal='right', vertical='center')     # F8 필드
ws_stock[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')     # G8 필드
ws_stock[f'H{row}'].alignment = Alignment(horizontal='right', vertical='center')     # H8 필드
ws_stock[f'I{row}'].alignment = Alignment(horizontal='right', vertical='center')     # I8 필드

# 04. 테두리 적용
ws_stock["A3"].border = thin_border   # A3 필드 -------->
ws_stock["B3"].border = thin_border   # B3 필드 
ws_stock["C3"].border = thin_border   # C3 필드 
ws_stock["D3"].border = thin_border   # D3 필드 
ws_stock["E3"].border = thin_border   # E3 필드
ws_stock["F3"].border = thin_border   # F3 필드 
ws_stock["G3"].border = thin_border   # G3 필드 
ws_stock["H3"].border = thin_border   # G3 필드 
ws_stock["I3"].border = thin_border   # I3 필드 
ws_stock["J3"].border = thin_border   # J3 필드(평가 손익 합계)
ws_stock["K3"].border = thin_border   # K3 필드(평가 손익 합계 Value) 
ws_stock["L3"].border = thin_border   # L3 필드(총평가 손익률(%))
ws_stock["M3"].border = thin_border   # M3 필드 (총평가 손익률(%) Value)
ws_stock["A4"].border = thin_border   # A4 필드(주식코드) Head  -------->
ws_stock["B4"].border = thin_border   # A4 필드
ws_stock["C4"].border = thin_border   # A4 필드
ws_stock["D4"].border = thin_border   # A4 필드
ws_stock["E4"].border = thin_border   # A4 필드
ws_stock["F4"].border = thin_border   # A4 필드
ws_stock["G4"].border = thin_border   # A4 필드
ws_stock["H4"].border = thin_border   # A4 필드
ws_stock["I4"].border = thin_border   # A4 필드
ws_stock["J4"].border = thin_border   # A4 필드
ws_stock["K4"].border = thin_border   # A4 필드
ws_stock["L4"].border = thin_border   # A4 필드
ws_stock["M4"].border = thin_border   # M4 필드(손절가)
ws_stock[f'A{row}'].border = thin_border    # A8 필드(합계)  -------->
ws_stock[f'B{row}'].border = thin_border    # B8 필드
ws_stock[f'C{row}'].border = thin_border    # C8 필드
ws_stock[f'D{row}'].border = thin_border    # D8 필드
ws_stock[f'E{row}'].border = thin_border   # E8 필드
ws_stock[f'F{row}'].border = thin_border   # F8 필드
ws_stock[f'G{row}'].border = thin_border    # G8 필드
ws_stock[f'H{row}'].border = thin_border   # H8 필드
ws_stock[f'I{row}'].border = thin_border   # I8 필드
ws_stock[f'J{row}'].border = thin_border   # J8 필드
ws_stock[f'K{row}'].border = thin_border   # K8 필드
ws_stock[f'L{row}'].border = thin_border   # L8 필드
ws_stock[f'M{row}'].border = thin_border   # M8 필드

# 05. 지정된 음영 색으로 음역 색칠하기(배경색 설정)
ws_stock["A3"].fill = blueDark3Fill      # A3 필드-------->
ws_stock["C3"].fill = grayDark2Fill      # C3 필드
ws_stock["D3"].fill = yellowFill         # D3 필드 
ws_stock["E3"].fill = grayDark2Fill      # E3 필드
ws_stock["H3"].fill = grayDark2Fill      # H3 필드
ws_stock["J3"].fill = grayDark2Fill      # J3 필드(평가 손익 합계)
ws_stock["L3"].fill = grayDark2Fill      # L3 필드(총평가 손익률(%))
ws_stock["A4"].fill = grayFill           # A4 필드(주식코드)
ws_stock["B4"].fill = grayFill   # A4 필드
ws_stock["C4"].fill = grayFill   # A4 필드
ws_stock["D4"].fill = grayFill   # A4 필드
ws_stock["E4"].fill = grayFill   # A4 필드
ws_stock["F4"].fill = grayFill   # A4 필드
ws_stock["G4"].fill = grayFill   # A4 필드
ws_stock["H4"].fill = grayFill   # A4 필드
ws_stock["I4"].fill = grayFill   # A4 필드
ws_stock["J4"].fill = grayFill   # A4 필드
ws_stock["K4"].fill = grayFill   # A4 필드
ws_stock["L4"].fill = grayFill   # A4 필드
ws_stock["M4"].fill = grayFill   # M4 필드(손절가) 
ws_stock[f'A{row}'].fill = orangeWeek2Fill      # A8 필드(합계)  -------->
ws_stock[f'B{row}'].fill = orangeWeek2Fill      # B8 필드
ws_stock[f'C{row}'].fill = orangeWeek2Fill      # C8 필드
ws_stock[f'D{row}'].fill = orangeWeek2Fill      # D8 필드
ws_stock[f'E{row}'].fill = orangeWeek2Fill      # E8 필드
ws_stock[f'F{row}'].fill = orangeWeek2Fill      # F8 필드
ws_stock[f'G{row}'].fill = orangeWeek2Fill      # G8 필드
ws_stock[f'H{row}'].fill = orangeWeek2Fill      # H8 필드
ws_stock[f'I{row}'].fill = orangeWeek2Fill      # I8 필드
ws_stock[f'J{row}'].fill = orangeWeek2Fill      # J8 필드
ws_stock[f'K{row}'].fill = orangeWeek2Fill      # K8 필드
ws_stock[f'L{row}'].fill = orangeWeek2Fill      # L8 필드
ws_stock[f'M{row}'].fill = orangeWeek2Fill       # M8 필드

print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_04_10] [주식 평가 금액]"+ str(evl_amt_sum) +"[신한은행 금액]"+ str(shinhanBValue) ) 
# ---------------------------------------------------------------------------------------------------------------------->
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


ws = wb["자산(2023)"]    # "자산(2023)" Sheet에 접근 @@@@@@@@ ===========>
prev_B7_amt = ws["B7"].value   # 70. [수협B] Sh 쑥쑥크는 아이적금(수종) 금액 ---> 이전 달 금액 
prev_B14_amt = ws["B14"].value   # 90. 총 합계 ---> 이전 달 금액
prev_A3 = ws["A3"].value   # 이전 자산 년월(예) 2023.09 자산)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_0] [일자(년월)_00]"+ str(prev_A3) )
prev_A3 = prev_A3[0:7]     # 이전 자산 년월(2023.08)

ws.insert_rows(3, 16)       # 월별 자산 제목(3번째 줄 위치에 16줄을 추가)
ws.merge_cells("A3:D3")     # A3 부터 D3까지 셀을 싱글 셀로 병합
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_1] [prev_B7_amt]"+ str(prev_B7_amt) +"[prev_B14_amt]"+ str(prev_B14_amt) )
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_2] [자산 년월]"+ str(astYYM) +"[오늘 년월]"+ str(now_ym) +"[일자(년월)]"+ str(prev_A3) )

# if str(astYYM) != str(now_ym) :   # 입력한 자산 년월와 오늘 년월이 다르면
#     result = pyautogui.alert("[자산(2023) Sheet] 입력한 자산 년월["+ str(astYYM) +"]와 오늘 년월["+ str(now_ym) +"]이 다릅니다.", title='[자산 년월 입력 오류]', button='OK')
#     sys.exit()    # 종료
# else :
#     if str(astYYM) == str(prev_A3) :   # 입력한 자산 년월와 이전 자산 년월이 같으면
#         result = pyautogui.alert("[자산(2023) Sheet] 입력한 자산 년월["+ str(astYYM) +"]이 이미 존재합니다.", title='[자산 년월 입력 오류]', button='OK')
#         sys.exit()    # 종료
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_3] [입력한 자산 년월]"+ str(astYYM) +"[오늘 년월]"+ str(now_ym) +"[이전 자산 년월]"+ str(prev_A3) ) 
# ---------------------------------------------------------------------------------------------------------------------->

astNm = "자산"     # 자산 명
clAcntCmt = ""     # A16 필드에 값 쓰기 --------> 결산 코멘트
# clAcntCmt = "결산 - 포항집에 김치 냉장고, 전자 레인지 구입(총 합계: 777,980원)"       # A16 필드에 값 쓰기 --------> 결산 코멘트
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■// [기초 Data 설정] //■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ----------->

H5_amt = shinhanBValue    # [2_Tab] 은행 결산@@ 신한은행[기초 Data] ■■
B6_amt = evl_amt_sum    # 77. 주식 투자 금액(평가 금액) ■■ -------->
C6_amt = stockPrchsAmt_sum   # 77. 주식 총 매입 금액
H6_amt = kakaoBValue     # [2_Tab] 은행 결산@@ 세이프박스(카카오 뱅크)) 금액[기초 Data] ■■
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_39] [B7_amt]"+ type(B7_amt) +"[prev_B7_amt]"+ type(prev_B7_amt) )

B7_amt = int(prev_B7_amt.replace(',','')) + 100000  # 70. [수협B] Sh 쑥쑥크는 아이적금(수종) 금액 -------->
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_5] [B7_amt]"+ str(B7_amt) +"[prev_B7_amt]"+ str(prev_B7_amt) )

H7_amt = shinhybBValue   # [2_Tab] 은행 결산@@ 신협 예금 금액[기초 Data] ■■
B8_amt = pensionIRP         # 91. 퇴직 연금(개인형 IPR) 금액[기초 Data] ■■ -------->
H8_amt = sokSavBValue     # [2_Tab] 은행 결산@@ OK 저축 은행 예금 금액

H9_amt = int(H5_amt) + int(H6_amt)  + int(H7_amt) + int(H8_amt)   # [06] 은행 예. 적금 총합 
B5_amt = H9_amt             # 01. 은행([06]) 금액
B9_amt = int(H9_amt) + int(B6_amt) + B7_amt + int(B8_amt)   # 1. 총 예적금 --------> 

B10_amt = insuSonValue   # 21. (무) 굿앤굿어린이CI보험(Hi1304)(현대해상, 진수종) 금액 -------->
B11_amt = insuValue      # 22. (무) 한아름행복플러스종합보험1404(한화손해보험, 잔태만) 금액 --------> 
B12_amt = int(B10_amt) + int(B11_amt)   # 2. 총 보험금 -------->
B13_amt = realtyValue    # 3. 부동산(현아트빌 404호) 금액 -------->

B14_amt = B9_amt + B12_amt + int(B13_amt)     # 90. 총 합계 금액 --------> 1. 총 예적금 + 2. 총 보험금 + 3. 부동산(현아트빌 404호) 
H12_amt = int(B14_amt) - int(prev_B14_amt.replace(',',''))   # [2_Tab] 전원 대비 증감@@ 총 합계 --> 90. 총 합계 - 이전 달 90. 총 합계
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_6] [총 합계]"+ str(H12_amt) +"[90. 총 합계_amt]"+ str(B14_amt) +"[이전 달 90. 총 합계]"+ str(prev_B14_amt) )

B15_amt = B9_amt            # 91. 가용 자산 금액 --------> 1. 총 예적금
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_7] [B15_amt]"+ str(B15_amt))
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->

ws.column_dimensions["G"].width = 18  # G열의 너비를 18로 설정

# 01. 값 쓰기
# ws.cell(row = 3, column = 1, value = astYYM +" "+ astNm)  # A3 필드에 값 쓰기(자산 년월)  -------->
ws["A3"].value = astYYM +" "+ astNm    # A3 필드에 값 쓰기(자산 년월)  -------->
ws["E3"].value = lastLastDt             # E3 필드에 값 쓰기
ws["G3"].value = '은행 결산'            # G3 필드에 값 쓰기
ws["J3"].value = '지출'                 # J3 필드에 값 쓰기
ws["A4"].value = '내용'                 # A4 필드에 값 쓰기  -------->
ws["B4"].value = '금액'                 # B4 필드에 값 쓰기
ws["C4"].value = '원금'                 # C4 필드에 값 쓰기
ws["D4"].value = '손익'                 # D4 필드에 값 쓰기
ws["E4"].value = '수익률(%)'            # E4 필드에 값 쓰기
ws["G4"].value = '은행 자산 상태'        # G4 필드에 값 쓰기
ws["H4"].value = '금액'                 # H4 필드에 값 쓰기
ws["J4"].value = '내역'                 # J4 필드에 값 쓰기
ws["K4"].value = '금액'                 # K4 필드에 값 쓰기
ws["A5"].value = '01. 은행 예. 적금([06])'   # A5 필드에 값 쓰기 -------->
ws["B5"].value = format(B5_amt, ',')        # B5 필드에 값 쓰기 ■■■
ws["C5"].value = ''   # C5 필드에 값 쓰기
ws["D5"].value = ''   # D5 필드에 값 쓰기
ws["E5"].value = ''   # E5 필드에 값 쓰기
ws["G5"].value = '신한은행'              # G5 필드에 값 쓰기
ws["H5"].value = format(H5_amt, ',')    # H5 필드에 값 쓰기
ws["J5"].value = '91. 월 고정 지출'      # J5 필드에 값 쓰기
ws["K5"].value = '1,766,187'   	        # K5 필드에 값 쓰기
ws["A6"].value = '77. 주식 투자'         # A6 필드에 값 쓰기 -------->
ws["B6"].value = format(B6_amt, ',')    # B6 필드에 값 쓰기 --> 77. 주식 투자 금액
ws["C6"].value = format(C6_amt, ',')    # C6 필드에 값 쓰기
ws["D6"].value = ''                     # D6 필드에 값 쓰기 
ws["G6"].value = '세이프박스(카카오 뱅크)'   # G6 필드에 값 쓰기
ws["H6"].value = format(H6_amt, ',')       # H6 필드에 값 쓰기 
ws["J6"].value = ''   # J6 필드에 값 쓰기
ws["K6"].value = ''   # K6 필드에 값 쓰기
ws["A7"].value = '70. [수협B] Sh 쑥쑥크는 아이적금(수종)'		# A7 필드에 값 쓰기 -------->
ws["B7"].value = format(B7_amt, ',')    # B7 필드에 값 쓰기 
ws["G7"].value = '신협 예금'             # G7 필드에 값 쓰기
ws["H7"].value = format(H7_amt, ',')    # H7 필드에 값 쓰기
ws["J7"].value = ''   # J7 필드에 값 쓰기
ws["K7"].value = ''   # K7 필드에 값 쓰기
ws["A8"].value = '91. 퇴직 연금(개인형 IPR)' 		 # A8 필드에 값 쓰기 -------->
ws["B8"].value = format(B8_amt, ',')    # B8 필드에 값 쓰기
ws["C8"].value = '16,000,000'   	    # C8 필드에 값 쓰기
ws["E8"].value = '100,000' 		        # E8 필드에 값 쓰기
ws["G8"].value = 'OK 저축 은행 예금'     # G8 필드에 값 쓰기
ws["H8"].value = format(H8_amt, ',')    # H8 필드에 값 쓰기
ws["J8"].value = ''   # J8 필드에 값 쓰기
ws["K8"].value = ''   # K8 필드에 값 쓰기
ws["A9"].value = '1. 총 예적금'  	        # A9 필드에 값 쓰기 -------->
ws["B9"].value = format(H9_amt, ',')       # B9 필드에 값 쓰기 
ws["G9"].value= '[06] 은행 예. 적금 총합'    # G9 필드에 값 쓰기
ws["H9"].value = format(H9_amt, ',')   	   # H9 필드에 값 쓰기 --> [06] 은행 예. 적금 총합 금액 
ws["J9"].value= ''   		# J9 필드에 값 쓰기
ws["K9"].value = ''   		# K9 필드에 값 쓰기
ws["A10"].value = '21. (무) 굿앤굿어린이CI보험(Hi1304)(현대해상, 진수종)'  		# A10 필드에 값 쓰기 -------->
ws["B10"].value = format(B10_amt, ',')     # B10 필드에 값 쓰기 
ws["J10"].value = ''   		# J10 필드에 값 쓰기
ws["K10"].value = ''   	    # K10 필드에 값 쓰기 
ws["A11"].value = '22. (무) 한아름행복플러스종합보험1404(한화손해보험, 잔태만)'	 # A11 필드에 값 쓰기 -------->
ws["B11"].value = format(B11_amt, ',')     # B11 필드에 값 쓰기
ws["G11"].value = '전원 대비 증감'  # G11 필드에 값 쓰기 
ws["H11"].value = '금액'         # H11 필드에 값 쓰기
ws["J11"].value = ''   			# J11 필드에 값 쓰기
ws["K11"].value = ''   			# K11 필드에 값 쓰기
ws["A12"].value = '2. 총 보험금' 			# A12 필드에 값 쓰기 --------> 
ws["B12"].value = format(B12_amt, ',')     # B12 필드에 값 쓰기 --> 2. 총 보험금 금액 
ws["G12"].value = '총 합계'   		        # G12 필드에 값 쓰기 
ws["H12"].value = format(H12_amt, ',')     # H12 필드에 값 쓰기
ws["J12"].value = '지출(A) 총합'   	        # J12 필드에 값 쓰기
ws["K12"].value = '1,766,187'   		   # K12 필드에 값 쓰기
ws["A13"].value = '3. 부동산(현아트빌 404호)' 	# A13 필드에 값 쓰기 -------->
ws["B13"].value = format(B13_amt, ',')        # B13 필드에 값 쓰기
ws["J13"].value = ''   # J13 필드에 값 쓰기
ws["K13"].value = ''   # K13 필드에 값 쓰기
ws["A14"].value = '90. 총 합계'  		    # A14 필드에 값 쓰기 -------->
ws["B14"].value = format(B14_amt, ',')     # B14 필드에 값 쓰기 --> 90. 총 합계 금액
ws["J14"].value = ''   # J8 필드에 값 쓰기
ws["K14"].value = ''   # K8 필드에 값 쓰기
ws["A15"].value = '91. 가용 자산' 		    # A15 필드에 값 쓰기 -------->
ws["B15"].value = format(B15_amt, ',')     # B15 필드에 값 쓰기 --> 91. 가용 자산 금액
ws["J15"].value = '수입(B) 총합'   	        # J15 필드에 값 쓰기
ws["K15"].value = '0'   		           # K15 필드에 값 쓰기 
ws["A16"].value = clAcntCmt                # A16 필드에 값 쓰기 --------> 결산 코멘트 
ws["J16"].value = '총 합계([B] - [A])'      # J16 필드에 값 쓰기
ws["K16"].value = '0'   		           # K16 필드에 값 쓰기 --> 총 합계([B] - [A]) 금액

# 02. 스타일 적용드 스타일 적용(글자 색은 빨갛게, 이탤릭, 두껍게 적용)
ws["A3"].font = Font(name="돋움", bold=True, size=12)   # A3 필드(자산 년월)  -------->
ws["E3"].font = Font(name="돋움", bold=True, size=9)    # E3 필드 
ws["G3"].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # G3 필드 --> 은행 결산
ws["J3"].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # J3 필드 --> 지출
ws["A4"].font = Font(name="돋움", bold=False, size=9)  # A4 필드 -------->
ws["B4"].font = Font(name="돋움", bold=False, size=9)  # B4 필드
ws["C4"].font = Font(name="돋움", bold=False, size=9)  # C4 필드
ws["D4"].font = Font(name="돋움", bold=False, size=9)  # D4 필드
ws["E4"].font = Font(name="돋움", bold=False, size=9)  # E4 필드 
ws["G4"].font = Font(name="돋움", bold=False, size=9)  # G4 필드 --> 은행 자산 상태
ws["H4"].font = Font(name="돋움", bold=False, size=9)  # H4 필드 --> 금액
ws["J4"].font = Font(name="돋움", bold=False, size=9)  # J4 필드 --> 내역
ws["K4"].font = Font(name="돋움", bold=False, size=9)  # K4 필드 --> 금액
ws["H4"].font = Font(name="돋움", bold=False, size=9)  # H4 필드 --> 금액
ws["A5"].font = Font(name="돋움", bold=False, size=9)  # A5 필드 -------->
ws["B5"].font = Font(name="돋움", bold=False, size=9, color="FF0000")  # B5 필드 --> 01. 은행 예. 적금([06]) 금액
ws["G5"].font = Font(name="돋움", bold=False, size=9)  # G5 필드 
ws["H5"].font = Font(name="돋움", bold=False, size=9)  # H5 필드 
ws["G5"].font = Font(name="돋움", bold=False, size=9)  # G5 필드 --> 91. 월 고정 지출
ws["H5"].font = Font(name="돋움", bold=False, size=9)  # H5 필드 --> 금액
ws["J5"].font = Font(name="돋움", bold=False, size=9)  # J5 필드 --> 91. 월 고정 지출
ws["K5"].font = Font(name="돋움", bold=False, size=9)  # K5 필드 --> 금액
ws["A6"].font = Font(name="돋움", bold=False, size=9)  # A6 필드 -------->
ws["B6"].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # B6 필드
ws["C6"].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # C6 필드 --> 77. 주식 총 매입 금액 
ws["G6"].font = Font(name="돋움", bold=False, size=9)  # G6 필드 
ws["H6"].font = Font(name="돋움", bold=False, size=9)  # H6 필드 
ws["A7"].font = Font(name="돋움", bold=False, size=9)  # A7 필드 --------> 
ws["B7"].font = Font(name="돋움", bold=False, size=9)  # B7 필드 
ws["G7"].font = Font(name="돋움", bold=False, size=9)  # G7 필드 
ws["H7"].font = Font(name="돋움", bold=False, size=9)  # H7 필드 
ws["A8"].font = Font(name="돋움", bold=False, size=9)  # A8 필드 -------->
ws["B8"].font = Font(name="돋움", bold=False, size=9)  # B8 필드
ws["C8"].font = Font(name="돋움", bold=False, size=9)  # C8 필드
ws["E8"].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # E8 필드
ws["G8"].font = Font(name="돋움", bold=False, size=9)  # G8 필드 
ws["H8"].font = Font(name="돋움", bold=False, size=9)  # H8 필드 
ws["A9"].font = Font(name="돋움", bold=True, size=9)    # A9 필드 -------->
ws["B9"].font = Font(name="돋움", bold=True, size=9)    # B9 필드
ws["G9"].font = Font(name="돋움", bold=False, size=9, color="FF0000")  # G9 필드
ws["H9"].font = Font(name="돋움", bold=False, size=9, color="FF0000")  # H9 필드
ws["A10"].font = Font(name="돋움", bold=False, size=9)  # A10 필드 -------->
ws["B10"].font = Font(name="돋움", bold=False, size=9)  # B10 필드
ws["A11"].font = Font(name="돋움", bold=False, size=9)  # A11 필드 -------->
ws["B11"].font = Font(name="돋움", bold=False, size=9)  # B11 필드
ws["G13"].font = Font(name="돋움", bold=False, size=9)  # G13 필드  
ws["H13"].font = Font(name="돋움", bold=False, size=9)  # H13 필드
ws["A12"].font = Font(name="돋움", bold=True, size=9)   # A12 필드 --------> 2. 총 보험금
ws["B12"].font = Font(name="돋움", bold=True, size=9)   # B12 필드
ws["G12"].font = Font(name="돋움", bold=True, size=9)   # G12 필드  
ws["H12"].font = Font(name="돋움", bold=True, size=9, color="FF0000")  # H12 필드 --> 2. 총 보험금 금액
ws["J12"].font = Font(name="돋움", bold=True, size=9)   # J12 필드 --> 지출(A) 총합
ws["K12"].font = Font(name="돋움", bold=True, size=9, color="FF0000")  # K12 필드 
ws["A13"].font = Font(name="돋움", bold=True, size=9)   # A13 필드 -------->
ws["B13"].font = Font(name="돋움", bold=True, size=9)   # B13 필드 
ws["A14"].font = Font(name="돋움", bold=True, size=9)   # A14 필드 -------->
ws["B14"].font = Font(name="돋움", bold=True, size=9, color="FF0000")  # B14 필드
ws["A15"].font = Font(name="돋움", bold=True, size=9)   # A15 필드 -------->
ws["B15"].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # B15 필드
ws["J15"].font = Font(name="돋움", bold=True, size=9)   # J15 필드 --> 수입(B) 총합
ws["K15"].font = Font(name="돋움", bold=True, size=9, color="FF0000")  # K15 필드
ws["A16"].font = Font(name="돋움", bold=False, size=9)  # A16 필드-------->
ws["J16"].font = Font(name="돋움", bold=True, size=9)  	# J16 필드 --> 총 합계([B] -[A])
ws["K16"].font = Font(name="돋움", bold=True, size=9, color="FF0000")  # K16 필드

# 03. alignment 적용
ws["A3"].alignment = Alignment(horizontal='center', vertical='center')      # A3 필드 alignment 설정 -------->
ws["E3"].alignment = Alignment(horizontal='right', vertical='center')       # E3 필드
ws["G3"].alignment = Alignment(horizontal='center', vertical='center')      # G3 필드 --> 은행 결산
ws["J3"].alignment = Alignment(horizontal='center', vertical='center')      # J3 필드 --> 지출
ws["A4"].alignment = Alignment(horizontal='center', vertical='center')      # A4 필드 -------->
ws["B4"].alignment = Alignment(horizontal='center', vertical='center')      # B4 필드
ws["C4"].alignment = Alignment(horizontal='center', vertical='center')      # C4 필드
ws["D4"].alignment = Alignment(horizontal='center', vertical='center')      # D4 필드
ws["E4"].alignment = Alignment(horizontal='center', vertical='center')      # E4 필드 
ws["G4"].alignment = Alignment(horizontal='center', vertical='center')      # G4 필드 --> 은행 자산 상태
ws["H4"].alignment = Alignment(horizontal='center', vertical='center')      # H4 필드 --> 금액
ws["J4"].alignment = Alignment(horizontal='center', vertical='center')      # J4 필드 --> 내역
ws["K4"].alignment = Alignment(horizontal='center', vertical='center')      # K4 필드 --> 금액
ws["A5"].alignment = Alignment(horizontal='left', vertical='center')        # A5 필드 -------->
ws["B5"].alignment = Alignment(horizontal='right', vertical='center')       # B5 필드 
ws["H5"].alignment = Alignment(horizontal='right', vertical='center')       # H5 필드  
ws["K5"].alignment = Alignment(horizontal='right', vertical='center')       # K5 필드 --> 금액
ws["A6"].alignment = Alignment(horizontal='left', vertical='center')        # A6 필드 -------->
ws["B6"].alignment = Alignment(horizontal='right', vertical='center')       # B6 필드 
ws["C6"].alignment = Alignment(horizontal='right', vertical='center')       # C6 필드 --> 77. 주식 총 매입 금액 
ws["H6"].alignment = Alignment(horizontal='right', vertical='center')       # H6 필드  
ws["B7"].alignment = Alignment(horizontal='right', vertical='center')       # B7 필드 --------> 
ws["H7"].alignment = Alignment(horizontal='right', vertical='center')       # H7 필드 
ws["B8"].alignment = Alignment(horizontal='right', vertical='center')       # B8 필드 -------->  
ws["C8"].alignment = Alignment(horizontal='right', vertical='center')       # C8 필드
ws["E8"].alignment = Alignment(horizontal='right', vertical='center')       # E8 필드  
ws["H8"].alignment = Alignment(horizontal='right', vertical='center')       # H8 필드 
ws["B9"].alignment = Alignment(horizontal='right', vertical='center')       # B9 필드 ------->
ws["H9"].alignment = Alignment(horizontal='right', vertical='center')       # H9 필드 
ws["B10"].alignment = Alignment(horizontal='right', vertical='center')      # B10 필드 ------->
ws["B11"].alignment = Alignment(horizontal='right', vertical='center')      # B11 필드 ------->
ws["G11"].alignment = Alignment(horizontal='center', vertical='center')     # G11 필드 ------->
ws["H11"].alignment = Alignment(horizontal='center', vertical='center')     # H11 필드
ws["B12"].alignment = Alignment(horizontal='right', vertical='center')      # B12 필드 --------> 금액(2. 총 보험금)
ws["G12"].alignment = Alignment(horizontal='center', vertical='center')     # G12 필드 
ws["H12"].alignment = Alignment(horizontal='right', vertical='center')      # H12 필드
ws["K12"].alignment = Alignment(horizontal='right', vertical='center')      # K12 필드 --> 금액(지출(A) 총합) 
ws["B13"].alignment = Alignment(horizontal='right', vertical='center')      # B13 필드 --------> 
ws["B14"].alignment = Alignment(horizontal='right', vertical='center')      # B14 필드 -------->
ws["B15"].alignment = Alignment(horizontal='right', vertical='center')      # B15 필드 -------->
ws["K15"].alignment = Alignment(horizontal='right', vertical='center')      # K15 필드 --> 금액(수입(B) 총합) 
ws["K16"].alignment = Alignment(horizontal='right', vertical='center')      # K16 필드 --> 금액(총 합계([B] -[A])) 

# 04. 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

ws["A3"].border = thin_border   # A3 필드 -------->
ws["B3"].border = thin_border   # B3 필드 
ws["C3"].border = thin_border   # C3 필드
ws["D3"].border = thin_border   # D3 필드 
ws["E3"].border = thin_border   # E3 필드
ws["G3"].border = thin_border   # G3 필드 --> image.png 
ws["H3"].border = thin_border   # H3 필드
ws["J3"].border = thin_border   # J3 필드 --> 지출 
ws["K3"].border = thin_border   # K3 필드  
ws["A4"].border = thin_border   # A4 필드 -------->
ws["B4"].border = thin_border   # B4 필드
ws["C4"].border = thin_border   # C4 필드
ws["D4"].border = thin_border   # D4 필드
ws["E4"].border = thin_border   # E4 필드
ws["G4"].border = thin_border   # G4 필드 --> 은행 자산 상태
ws["H4"].border = thin_border   # H4 필드 --> 금액
ws["J4"].border = thin_border   # J4 필드 --> 내역
ws["K4"].border = thin_border   # K4 필드 --> 금액
ws["A5"].border = thin_border   # A5 필드 -------->
ws["B5"].border = thin_border   # B5 필드
ws["C5"].border = thin_border   # C5 필드
ws["D5"].border = thin_border   # D5 필드
ws["E5"].border = thin_border   # E5 필드
ws["G5"].border = thin_border   # G5 필드
ws["H5"].border = thin_border   # H5 필드
ws["J5"].border = thin_border   # J5 필드 --> 91. 월 고정 지출
ws["K5"].border = thin_border   # K5 필드 --> 금액 
ws["A6"].border = thin_border   # A6 필드 -------->
ws["B6"].border = thin_border   # B6 필드
ws["C6"].border = thin_border   # C6 필드
ws["D6"].border = thin_border   # D6 필드 
ws["E6"].border = thin_border   # E6 필드
ws["G6"].border = thin_border   # G6 필드
ws["H6"].border = thin_border   # H6 필드
ws["J6"].border = thin_border   # J6 필드 --> 빈 칸
ws["K6"].border = thin_border   # K6 필드 --> 빈 칸 
ws["A7"].border = thin_border   # A7 필드 -------->
ws["B7"].border = thin_border   # B7 필드
ws["C7"].border = thin_border   # C7 필드
ws["D7"].border = thin_border   # D7 필드
ws["E7"].border = thin_border   # E7 필드
ws["G7"].border = thin_border   # G7 필드
ws["H7"].border = thin_border   # H7 필드
ws["J7"].border = thin_border   # J7 필드 --> 빈 칸
ws["K7"].border = thin_border   # K7 필드 --> 빈 칸 
ws["A8"].border = thin_border   # A8 필드 -------->
ws["B8"].border = thin_border   # B8 필드
ws["C8"].border = thin_border   # C8 필드
ws["D8"].border = thin_border   # D8 필드
ws["E8"].border = thin_border   # E8 필드
ws["G8"].border = thin_border   # G8 필드
ws["H8"].border = thin_border   # H8 필드
ws["J8"].border = thin_border   # J8 필드 --> 빈 칸
ws["K8"].border = thin_border   # K8 필드 --> 빈 칸 
ws["A9"].border = thin_border   # A9 필드 -------->
ws["B9"].border = thin_border   # B9 필드
ws["C9"].border = thin_border   # C9 필드
ws["D9"].border = thin_border   # D9 필드
ws["E9"].border = thin_border   # E9 필드
ws["G9"].border = thin_border   # G9 필드
ws["H9"].border = thin_border   # H9 필드
ws["J9"].border = thin_border   # J9 필드 --> 빈 칸
ws["K9"].border = thin_border   # K9 필드 --> 빈 칸 
ws["A10"].border = thin_border  # A10 필드 -------->
ws["B10"].border = thin_border  # B10 필드
ws["C10"].border = thin_border  # C10 필드
ws["D10"].border = thin_border  # D10 필드
ws["E10"].border = thin_border  # E10 필드
ws["H10"].border = thin_border  # H10 필드
ws["J10"].border = thin_border  # J10 필드 --> 빈 칸
ws["K10"].border = thin_border  # K10 필드 --> 빈 칸 
ws["A11"].border = thin_border  # A11 필드 -------->
ws["B11"].border = thin_border  # B11 필드
ws["C11"].border = thin_border  # C11 필드
ws["D11"].border = thin_border  # D11 필드
ws["E11"].border = thin_border  # E11 필드
ws["G11"].border = thin_border  # G11 필드
ws["H11"].border = thin_border  # H11 필드
ws["J11"].border = thin_border  # J11 필드 --> 빈 칸
ws["K11"].border = thin_border  # K11 필드 --> 빈 칸 
ws["A12"].border = thin_border  # A12 필드 -------->
ws["B12"].border = thin_border  # B12 필드
ws["C12"].border = thin_border  # C12 필드
ws["D12"].border = thin_border  # D12 필드
ws["E12"].border = thin_border  # E12 필드
ws["G12"].border = thin_border  # G12 필드
ws["H12"].border = thin_border  # H12 필드
ws["J12"].border = thin_border 	# J12 필드 ----> 지출(A) 총합
ws["K12"].border = thin_border  # K12 필드 
ws["A13"].border = thin_border  # A13 필드 -------->
ws["B13"].border = thin_border  # B13 필드
ws["C13"].border = thin_border  # C13 필드
ws["D13"].border = thin_border  # D13 필드
ws["E13"].border = thin_border  # E13 필드
ws["J13"].border = thin_border  # J13 필드 --> 빈 칸
ws["K13"].border = thin_border  # K13 필드 --> 빈 칸 
ws["A14"].border = thin_border  # A14 필드 -------->
ws["B14"].border = thin_border  # B14 필드
ws["C14"].border = thin_border  # C14 필드
ws["D14"].border = thin_border  # D14 필드
ws["E14"].border = thin_border  # E14 필드
ws["J14"].border = thin_border  # J14 필드 --> 빈 칸
ws["K14"].border = thin_border  # K14 필드 --> 빈 칸 
ws["A15"].border = thin_border  # A15 필드 -------->
ws["B15"].border = thin_border  # B15 필드
ws["C15"].border = thin_border  # C15 필드
ws["D15"].border = thin_border  # D15 필드
ws["E15"].border = thin_border  # E15 필드
ws["J15"].border = thin_border  # J15 필드 ----> 지출(A) 총합
ws["K15"].border = thin_border  # K15 필드 
ws["A16"].border = thin_border  # A16 필드 -------->
ws["B16"].border = thin_border  # B16 필드
ws["C16"].border = thin_border  # C16 필드
ws["D16"].border = thin_border  # D16 필드
ws["E16"].border = thin_border  # E16 필드
ws["J16"].border = thin_border 	# J16 필드 ----> 총 합계([B] -[A])
ws["K16"].border = thin_border  # K16 필드 

# 05. 지정된 음영 색으로 음역 색칠하기(배경색 설정)
ws["A3"].fill = orangeFill      # A3 필드-------->
ws["E3"].fill = orangeFill      # E3 필드
ws["G3"].fill = blueDark2Fill   # G3 필드 --> 은행 결산 
ws["J3"].fill = blueDark2Fill   # J3 필드 --> 지출 
ws["A4"].fill = grayFill        # A4 필드 -------->
ws["B4"].fill = grayFill        # B4 필드
ws["C4"].fill = grayFill        # C4 필드
ws["D4"].fill = grayFill        # D4 필드
ws["E4"].fill = grayFill        # E4 필드 
ws["G4"].fill = grayFill        # G4 필드 --> 은행 자산 상태
ws["H4"].fill = grayFill        # H4 필드 --> 금액
ws["J4"].fill = grayFill        # J4 필드 --> 내역
ws["K4"].fill = grayFill        # K4 필드 --> 금액
ws["A5"].fill = blueFill        # A5 필드 -------->
ws["B5"].fill = blueFill        # B5 필드
ws["C5"].fill = blueFill        # C5 필드
ws["D5"].fill = blueFill        # D5 필드
ws["E5"].fill = blueFill        # E5 필드 
ws["H5"].fill = blueFill        # H5 필드 
ws["B6"].fill = yellowFill      # B6 필드 -------->
ws["H6"].fill = blueFill        # H6 필드
ws["B7"].fill = blueDarkFill    # B7 필드 --------> 
ws["H7"].fill = blueFill        # H7 필드
ws["B8"].fill = blueDarkFill    # B8 필드 -------->
ws["E8"].fill = violetFill      # E8 필드
ws["H8"].fill = blueFill        # H8 필드 
ws["A9"].fill = greenWeekFill   # A9 필드 -------->
ws["B9"].fill = greenWeekFill   # B9 필드
ws["C9"].fill = greenWeekFill   # C9 필드
ws["D9"].fill = greenWeekFill   # D9 필드
ws["E9"].fill = greenWeekFill   # E9 필드 
ws["G9"].fill = blueDark2Fill   # G9 필드 
ws["H9"].fill = blueDark2Fill   # H9 필드 
ws["G11"].fill = grayFill       # G11 필드
ws["H11"].fill = grayFill       # H11 필드
ws["A12"].fill = greenFill      # A12 필드 -------->
ws["B12"].fill = greenFill      # B12 필드
ws["C12"].fill = greenFill      # C12 필드
ws["D12"].fill = greenFill      # D12 필드
ws["E12"].fill = greenFill      # E12 필드
ws["G12"].fill = yellowFill     # G12 필드
ws["H12"].fill = yellowFill     # H12 필드
ws["J12"].fill  = greenFill     # J12 필드 --> 지출(A) 총합
ws["K12"].fill = greenFill      # K12 필드
ws["A13"].fill = greenFill      # A13 필드 -------->
ws["B13"].fill = greenFill      # B13 필드
ws["C13"].fill = greenFill      # C13 필드
ws["D13"].fill = greenFill      # D13 필드
ws["E13"].fill = greenFill      # E13 필드 
ws["A14"].fill = yellowFill     # A14 필드 -------->
ws["B14"].fill = yellowFill     # B14 필드
ws["C14"].fill = yellowFill     # C14 필드
ws["D14"].fill = yellowFill     # D14 필드
ws["E14"].fill = yellowFill     # E14 필드
ws["A15"].fill = yellowFill     # A15 필드 -------->
ws["B15"].fill = yellowFill     # B15 필드
ws["C15"].fill = yellowFill     # C15 필드
ws["D15"].fill = yellowFill     # D15 필드
ws["E15"].fill = yellowFill     # E15 필드
ws["J15"].fill = grayDarkFill   # J15 필드 --> 수입(B) 총합
ws["K15"].fill = grayDarkFill   # K15 필드
ws["A16"].fill = orangeWeekFill # A16 필드 -------->
ws["B16"].fill = orangeWeekFill # B16 필드
ws["C16"].fill = orangeWeekFill # C16 필드
ws["D16"].fill = orangeWeekFill # D16 필드
ws["E16"].fill = orangeWeekFill # E16 필드
ws["J16"].fill = yellowFill     # J16 필드 --> 총 합계([B] -[A])
ws["K16"].fill = yellowFill     # K16 필드
# ===============================================================================================================================


# 90 점 넘는 셀에 대해서 초록색으로 적용
# for row in ws.rows:
#     for cell in row:
#         # if row.column == 1 and cell.column == 1: # A 번호열은 제외       
#         #     cell.alignment = Alignment(horizontal="center", vertical="center")   # 각 cell 에 대해서 정렬
#         # else:
#         #     cell.alignment = Alignment(horizontal="center", vertical="center")   # 각 cell 에 대해서 정렬   
#         #     # center, left, right, top, bottom

#         if cell.column == 1: # A 번호열은 제외
#             continue

#         # cell 이 정수형 데이터이고 90 점보다 높으면
#         # if isinstance(cell.value, int) and cell.value > 90:
#         #     cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로 설정
#         #     cell.font = Font(color="FF0000") # 폰트 색상 변경
# ===============================================================================================================================

# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정

rsltFileNm = "01. 자산 검증("+ astYM +")_rslt.xlsx"    # 결과 파일 명(01. 자산 검증(23.08)_rslt.xlsx)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_91] [결과 파일 명]"+ rsltFileNm )

wb.save(urlPath + rsltFileNm)
# wb.save("01. 자산 검증(23.08)_rslt.xlsx")
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_99] ■■■■■■ [######################### [자산 검증 파일 TEST End] #########################] ■■■■■■\n\n\n\n")