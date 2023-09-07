# ! /53. stock.py		# 주식 파일 

from selenium import webdriver   # 웹드라이버 Lib
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager   # 웹드라이버 매니저 Lib ==> # 주식에서  
import csv		# CSV 파일 Lib 
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook # 파일 불러오기 --> [기초 Data]  

print("[@_T] ■■■ [/53. stock.py] ==> [T_01] ■■■■■■ [######################### [쇼핑몰 TEST Start] #########################] ■■■■■■ ")

options = Options() 
options.add_experimental_option("detach", True)  # 브라우저 바로 닫힘 방지
options.add_experimental_option("excludeSwitches", ["enable-logging"])  # 불필요한 메시지 제거 
# options.add_experimental_option("excludeSwitches", ["enable-automation"])	# 크롬이 자동화된 테스트 소프트웨어에 의해 제어 알람 제거

driver = webdriver.Chrome(options=options)
print("[@_T] ■■■ [/53. stock.py] ==> [T_02]" ) 

urlPath = "rpa_basic/3_web/"   # URL 경로  D:\PythonWorkspace\rpa_basic\3_web\
openFileNm = "53. stock Data.xlsx"   # 오픈 파일 명(01. 자산 검증(23.08).xlsx) 
wb = load_workbook(urlPath + openFileNm, data_only=True)    # 오픈 파일을 wb을 불러옴(data_only=True: 수식이 아닌 실제 데이터를 가지고 옴)
print("[@_T] ■■■ [/53. stock.py] ==> [T_03]" ) 

ws = wb.active  # 현재 활성화된 Sheet에 접근

# 종목 코드 리스트
codes = [
    '005380'   # 현대차     
    , '068270' # 셀트리온
    , '096770'  # SK이노베이션
]

row = 0 
print("[@_T] ■■■ [/53. stock.py] ==> [T_05]" )  

for code in codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"    # 네이버 증권 종목(f: 포맷 문자열, 리터럴 사용)
    # url = "https://finance.naver.com/item/sise.naver?code="+ code    # 네이버 증권 종목     
    # url = 'https://finance.naver.com/item/sise.naver?code=005380'    # 네이버 증권 종목 ===> TEST @@@ ===>

    response = requests.get(url)     # 웹 사이트 열기

    if response.status_code == 200:        
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        price = soup.select_one("#_nowVal").text    # 현재가 ==>sCSS_SELECTOR
        print("[@_T] ■■■ [/53. stock.py] ==> [T_50] [row_번째]"+ str(price) +"[현재가]"+ str(price) +"[url_주소]"+ str(url) )  

        price = price.replace(',', '')    
        # price =  5000000    # 현재가  TEST @@@ ===>
        # print("[@_T] ■■■ [/53. stock.py] ==> [T_50] [row_번째] [soup]"+ str(soup) )

    else : 
        print(response.status_code) 
    print("[@_T] ■■■ [/53. stock.py] ==> [T_51] [row_번째]"+ str(row) +"[종목 코드]"+ code +"[현재가]"+ str(price) ) 

    ws[f'C{row+3}'] = int(price)    # 현재가(엑셀)
    row = row + 1
print("[@_T] ■■■ [/53. stock.py] ==> [T_60]" )

wb.save(urlPath + openFileNm)

time.sleep(90)   # 10초 대기
print("[@_T] ■■■ [/53. stock.py] ==> [T_90]" )

# driver.quit()

print("[@_T] ■■■ [/53. stock.py] ==> [T_99] ■■■■■■ [######################### [쇼핑몰 TEST End] #########################] ■■■■■■\n\n\n\n")
