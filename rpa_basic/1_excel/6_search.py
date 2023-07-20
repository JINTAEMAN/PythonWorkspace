from openpyxl import load_workbook
<<<<<<< HEAD

wb = load_workbook("sample.xlsx")
ws = wb.active
print("\n\n\n\n[@_T] ■■■ [/6_search.py] ==> [T_01] ■■■■■■ [######################### [파일 열기 TEST Start] #########################] ■■■■■■ ")
=======
wb = load_workbook("sample.xlsx")
ws = wb.active
>>>>>>> eda17eae7e526c9fb1f71b53e863ca50f6374b75

for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

<<<<<<< HEAD
wb.save("sample_modified.xlsx")
print("\n\n[@_T] ■■■ [/6_search.py] ==> [T_01] ■■■■■■ [######################### [파일 열기 TEST End] #########################] ■■■■■■\n\n\n\n")
=======
wb.save("sample_modified.xlsx")
>>>>>>> eda17eae7e526c9fb1f71b53e863ca50f6374b75
