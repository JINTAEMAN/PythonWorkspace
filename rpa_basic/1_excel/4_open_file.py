from openpyxl import load_workbook # 파일 불러오기

urlPath =  "rpa_basic/1_excel/"   # URL 경로
openFileNm = "sample.xlsx"   # 오픈 파일 명(sample.xlsx)
wb = load_workbook(urlPath + openFileNm)   # 오픈 파일을 wb을 불러옴 
ws = wb.active  # 활성화 되어있는 시트 설정
print("\n\n\n\n[@_TT] ■■■ [/4_open_file.py] ==> [T_01] ■■■■■■ [######################### [파일 열기 TEST Start] #########################] ■■■■■■ ")

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
#     print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
    print()


print("\n\n[@_TT] ■■■ [/4_open_file.py] ==> [T_01] ■■■■■■ [######################### [파일 열기 TEST End] #########################] ■■■■■■\n\n\n\n")