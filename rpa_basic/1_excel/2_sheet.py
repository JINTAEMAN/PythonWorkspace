# 2_sheet.py
from openpyxl import Workbook

wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경
# print("\n\n[@_T] ■ [/2_sheet.py] ==> [T_40] [sheet 명]"+ ws.title +"[Tab 배경색]"+ ws.sheet_properties.tabColor )
print("\n\n[@_T] ■ [/2_sheet.py] ==> [T_40] [sheet 명]"+ ws.title )

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index 에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet 에 접근

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"
print("\n\n[@_T] ■ [/2_sheet.py] ==> [T_50] [target sheet 명]"+ target.title )

<<<<<<< HEAD
wb.save("sample.xlsx")
=======
wb.save("sample.xlsx")


>>>>>>> eda17eae7e526c9fb1f71b53e863ca50f6374b75
