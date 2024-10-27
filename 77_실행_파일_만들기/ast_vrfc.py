# BeautifulSouprequests# -*- coding: utf-8 -*-	    # 문자 인코딩(한글 사용) 
# ! /ast_vrfc.py		    # 자산 검증 파일 

import shelve
from openpyxl import load_workbook # 파일 불러오기 --> [기초 Data]
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles import PatternFill
import os       # 파이씬 os 제어 Lib
import sys	    # 파이씬 인터프리터 제어 Lib
import pyautogui    # 마우스와 키보드 제어 Lib
import time     # 시간 Lib
from datetime import datetime   # datetime Lib
import datetime

import calendar  #   calendar Lib
# import requests
# from bs4 import BeautifulSoup

urlPath = ""   # 01. URL 경로(엑셀 폴더 경로)
# urlPath = "01_ast_vrfc/"   # 01. URL 경로(엑셀 폴더 경로)  # ■■■■■■■ ===> TEST 수행(테스트용) @@@ ===>    
opeParaFileNm = "01_1. paramYM.txt"   # 02. 파라 파일명 # ■■■■■■■ ===> Real
# opeParaFileNm = "01_1. paramYM_24.12.txt"   # 02. 파라 파일명(24.12) --> 년말 테스트  # ■■■■■■■ ===> TEST @@@  ===>

# /**
# * @description 파일에서 파라미터 가져오는 함수()
# */C
def readParameters():
    #  print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_01]")
    print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_31] [URL 경로]"+ str(urlPath) +"[파라 파일명]"+ str(opeParaFileNm) )

    file = open(urlPath + opeParaFileNm, 'rt', encoding='utf-8-sig')	# properties.txt 파일 내용]
    print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_51] [file]"+ str(file) )

    parameters = file.read().split(";")   # 자산 년월 parameters[01. 현재 년월[입력]]
    print("[@_T] ■■■ [/ast_vrfc.py] [readParameters]==> [T_92] [01. 이전 년월[입력]]"+ str(parameters[0]) +"02. 출력할 년월[현재 월]■]"+ str(parameters[1] ) ) 

    return parameters
# ---------------------------------------------------------------------------------------------------------------------->


# 00. 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 00. 음영 색 지정
orangeFill = PatternFill(start_color='F79646', end_color='F79646', fill_type='solid')       # 오렌지 색
orangeWeekFill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')   # 연한 오렌지
orangeWeek2Fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')  # 연한 오렌지2
orangeWeek3Fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')  # 연한 오렌지3 
grayFill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')         # 회 색
grayDarkFill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')     # 진한 회색 
grayDark2Fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')    # 진한 회색2  
blueFill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')         # 하늘 색
blueDarkFill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')     # 진한 하늘 색
blueDark2Fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')    # 진한 하늘 색2
blueDark3Fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')    # 진한 파란 색
yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')       # 노란 색
violetFill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')       # 보라 색
greenWeekFill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')    # 연한 녹색
greenFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')        # 녹색
whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')        # 흰 색
# ---------------------------------------------------------------------------------------------------------------------->
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


print("\n\n [@_T] ■■■ [/ast_vrfc.py] ==> [T_01] ■■■■■■ [######################### [자산 검증 파일 TEST Start] #########################] ■■■■■■ ")
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_01_01]" )

now_ym = str(datetime.datetime.now()).replace('-','.')     # 오늘 년월(년.월) (2023.08.31)
dt_now = datetime.datetime.now() 
lastLastDt = dt_now.strftime("%Y") +"."+ dt_now.strftime("%m") +"."+ dt_now.strftime("%d")   # 자산 마지막 일 ■■■■■■ "2023.08.31"

parameters = readParameters()      # 파일에서 파라미터 가져오는 함수() --> [01. 현재 년월[입력]]
input_astYYYYMM = str(parameters[0])     # 자산 년월 --> 이전 년월
input_astYYYYMM = input_astYYYYMM.split(".")     
input_astYY = str(input_astYYYYMM[0][2:4])  # 자산 년[입력] -->ㅎㄵ 년월
input_astMM = int(input_astYYYYMM[1])    # 자산 년월[입력]
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_02] [01. 현재 년월[입력][★]]"+ str(parameters[0]) +"[02. 출력할 년월■]"+ str(parameters[1]) +"[3. 오늘 년월(오늘 년월■)]"+ str(input_astMM) ) 

if str(parameters[0]) == None :   # 자산 년월 미입력 이면
    result = pyautogui.alert("자산 년월을 입력하세요. 예) 2023.08", title='[자산 년월 입력 오류]', button='OK')
    sys.exit()    # 종료

astYYM = parameters[1][0:7]    # 자산 년월[입력] --> 현재 년월(2023.08: 6자리)
astYM = parameters[1][2:7]  # 자산 년월(23.08: 4자리) 
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_04] [01. 자산 년월(현재 년월)[★]]"+ str(astYYM) +"[02. 자산 년월(현재 년월)■]"+ str(astYM) ) 

openFileNm = "02. 자산 검증("+ str(input_astYY) +"."+ format(input_astMM, '02') +").xlsx"   # 오픈 파일 명(02. 자산 검증(23.08).xlsx) 
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_04] [URL 경로]"+ str(urlPath) +"[오픈 파일]"+ str(openFileNm) )

wb = load_workbook(urlPath + openFileNm, data_only=True)    # 오픈 파일을 wb을 불러옴(data_only=True: 수식이 아닌 실제 데이터를 가지고 옴)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_02_05] [URL 경로]"+ str(urlPath) +"[오픈 파일]"+ str(openFileNm) +"[wb]"+ str(wb) ) 
# ---------------------------------------------------------------------------------------------------------------------->
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


print("\n\n [@_T] ■■■ [/ast_vrfc.py] ==> [T_01] ■■■■■■ [######################### [11. 계좌별잔고 Tab(기준 정보) 작업 Start] #########################] ■■■■■■ ")

ws_01_actBalnc = wb["계좌별잔고"]   # "계좌별잔고" Sheet(Tab)에 접근 @@@@@@@@ ===========> 
totAstYear = "02. 총자산("+ str(astYYM[2:4]) +"Y)"
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_11_01] [계좌별잔고 Tab Sheet(Tab)에 접근 @@@@@@@@ ===========>]")
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_11_02] [00. totAstYear]"+ str(totAstYear) +"[자산 년]"+ str(input_astYYYYMM[0]) +"[자산 월]"+ str(input_astYYYYMM[1]) )

add_row = 0     # 줄 추가
rowNo = add_row + 16

# 01. 주식 월별 보유 현황[계좌별잔고 Tab] @@@
posses_qty1 = ws_01_actBalnc[f'B{rowNo}'].value  	   # 01. 보유 수량(현대차) DDDD
cur_amt1 = ws_01_actBalnc[f'C{rowNo}'].value          # 02. 현재가(현대차)
ave_prchs_amt1 = ws_01_actBalnc[f'D{rowNo}'].value    # 03. 평균매입가(현대차) DDDD
prchs_amt1 = ws_01_actBalnc[f'E{rowNo}'].value        # 04. 매입금액(현대차)   DDDD
eval_amt1 = ws_01_actBalnc[f'F{rowNo}'].value         # 05. 평가금액(현대차)
eval_profit_loss1 = ws_01_actBalnc[f'G{rowNo}'].value   # 06. 평가손익(현대차)
eval_profit_loss_rate1 = ws_01_actBalnc[f'H{rowNo}'].value    # 07. 손익률(%)(현대차)
eval_profit_loss_rate1 = round((eval_profit_loss_rate1 * 100), 2)  
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_11_03] [(현대차)_손익률(%)]"+ str(eval_profit_loss_rate1) )  # d = round(1.23456, 2)

rowNo = add_row + 17
posses_qty2 = ws_01_actBalnc[f'B{rowNo}'].value  	   # 01. 보유 수량(셀트리온)
cur_amt2 = ws_01_actBalnc[f'C{rowNo}'].value          # 02. 현재가(셀트리온)
ave_prchs_amt2 = ws_01_actBalnc[f'D{rowNo}'].value    # 03. 평균매입가(셀트리온)
prchs_amt2 = ws_01_actBalnc[f'E{rowNo}'].value        # 04. 매입금액(셀트리온) 
eval_amt2 = ws_01_actBalnc[f'F{rowNo}'].value         # 05. 평가금액(셀트리온)
eval_profit_loss2 = ws_01_actBalnc[f'G{rowNo}'].value   # 06. 평가손익(셀트리온)
eval_profit_loss_rate2 = ws_01_actBalnc[f'H{rowNo}'].value    # 07. 손익률(%)(셀트리온)
eval_profit_loss_rate2 = round((eval_profit_loss_rate2 * 100), 2)

rowNo = add_row + 18
posses_qty3 = ws_01_actBalnc[f'B{rowNo}'].value  	  # 01. 보유 수량(SK이노베이션)
cur_amt3 = ws_01_actBalnc[f'C{rowNo}'].value          # 02. 현재가(SK이노베이션)
ave_prchs_amt3 = ws_01_actBalnc[f'D{rowNo}'].value    # 03. 평균매입가(SK이노베이션)
prchs_amt3 = ws_01_actBalnc[f'E{rowNo}'].value        # 04. 매입금액(SK이노베이션) 
eval_amt3 = ws_01_actBalnc[f'F{rowNo}'].value         # 05. 평가금액(SK이노베이션)
eval_profit_loss3 = ws_01_actBalnc[f'G{rowNo}'].value   # 06. 평가손익(SK이노베이션)
eval_profit_loss_rate3 = ws_01_actBalnc[f'H{rowNo}'].value    # 07. 손익률(%)(SK이노베이션)
eval_profit_loss_rate3 = round((eval_profit_loss_rate3 * 100), 2)

# rowNo = add_row + 19
# prchs_amt4 = ws_01_actBalnc[f'E{rowNo}'].value        # 04. 매입금액(02. 발행어음) 
# eval_amt4 = ws_01_actBalnc[f'F{rowNo}'].value         # 05. 평가금액(02. 발행어음)
# eval_profit_loss4 = ws_01_actBalnc[f'G{rowNo}'].value   # 06. 평가손익(02. 발행어음) 
# # ===============================================================================================================================
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


print("\n\n [@_T] ■■■ [/ast_vrfc.py] ==> [T_01] ■■■■■■ [######################### [21. 퇴직신탁 Tab(기준 정보) 작업 Start] #########################] ■■■■■■ ")

ws_02_irp = wb["퇴직신탁"]   # "퇴직신탁" Sheet(Tab)에 접근 @@@@@@@@ ===========>
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_01] [퇴직신탁 Tab Sheet(Tab)에 접근 @@@@@@@@ ===========>]")

# 퇴직신탁 보유현황(IRP) @@@ 
add_row = 0     # 줄 추가
rowNo = add_row + 6
irp_stock_nm_1 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_1(퇴직 연금)
irp_eval_amt_1 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_1(퇴직 연금)
irp_eval_profit_loss_1 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_1(퇴직 연금)
irp_eval_profit_loss_rate_1 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_1(퇴직 연금)
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_02] [평가금액_1(퇴직 연금)]"+ str(irp_eval_amt_1) +"[평가손익_1(퇴직 연금)]"+ str(irp_eval_profit_loss_1) +"[손익률(%)_1(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

irp_eval_amt_1 = irp_eval_amt_1.split("원")   ## 평가금액_1(퇴직 연금)
irp_eval_amt_1 =irp_eval_amt_1[0]

irp_eval_profit_loss_rate_1 = irp_eval_profit_loss_rate_1.split("%")   # 손익률(%)_1(퇴직 연금)
irp_eval_profit_loss_rate_1 = irp_eval_profit_loss_rate_1[0]
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_4] 손익률(%)_1퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

irp_prchs_amt = ws_02_irp[f'L{rowNo}'].value  	   # 매입금액 총합(퇴직 연금) ==> 03. IRP 월별 보유 현황) @@@
irp_eval_amt = ws_02_irp[f'M{rowNo}'].value  	   # 평가금액 총합(퇴직 연금) ==> 03. IRP 월별 보유 현황) @@@
irp_eval_profit_loss = ws_02_irp[f'N{rowNo}'].value  	   # 평가손익 총합(퇴직 연금) ==> 03. IRP 월별 보유 현황) @@@
irp_eval_profit_loss_rate = ws_02_irp[f'O{rowNo}'].value     # 손익률(%) 총합(퇴직 연금) ==> 03. IRP 월별 보유 현황) @@@
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_6] [매입금액(퇴직 연금)]"+ str(irp_prchs_amt) +"[평가금액(퇴직 연금)]"+ str(irp_eval_amt) +"[평가손익(퇴직 연금)]"+ str(irp_eval_profit_loss) +"[손익률(퇴직 연금)]"+ str(irp_eval_profit_loss_rate) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 7
irp_stock_nm_2 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_2(퇴직 연금)
irp_eval_amt_2 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_2(퇴직 연금)
irp_eval_profit_loss_2 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_2(퇴직 연금)
irp_eval_profit_loss_rate_2 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_2(퇴직 연금)
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_7] [평가금액__2(퇴직 연금)]"+ str(irp_eval_amt_2) +"[평가손익__2(퇴직 연금)]"+ str(irp_eval_profit_loss_2) +"[손익률(%)__2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_2) )

irp_eval_amt_2 = irp_eval_amt_2.split("원")   ## 평가금액_2(퇴직 연금)
irp_eval_amt_2 =irp_eval_amt_2[0]
irp_eval_profit_loss_rate_2 = irp_eval_profit_loss_rate_2.split("%")   # 손익률(%)_2(퇴직 연금)
irp_eval_profit_loss_rate_2 = irp_eval_profit_loss_rate_2[0]  
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_8] [손익률(%)__2퇴직 연금)]"+ str(irp_eval_profit_loss_rate_2) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 8
irp_stock_nm_3 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_3(퇴직 연금)
irp_eval_amt_3 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_3(퇴직 연금)
irp_eval_profit_loss_3 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_3(퇴직 연금)
irp_eval_profit_loss_rate_3 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_3(퇴직 연금)
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_9] [평가금액__3(퇴직 연금)]"+ str(irp_eval_amt_3) +"[평가손익__3(퇴직 연금)]"+ str(irp_eval_profit_loss_3) +"[손익률(%)__3(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_3) )

irp_eval_amt_3 = irp_eval_amt_3.split("원")   ## 평가금액_3(퇴직 연금)
irp_eval_amt_3 =irp_eval_amt_3[0]
irp_eval_profit_loss_rate_3 = irp_eval_profit_loss_rate_3.split("%")   # 손익률(%)_3(퇴직 연금)
irp_eval_profit_loss_rate_3 = irp_eval_profit_loss_rate_3[0]
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 9 
irp_stock_nm_4 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_4(퇴직 연금)
irp_eval_amt_4 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_4(퇴직 연금)
irp_eval_profit_loss_4 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_4(퇴직 연금)
irp_eval_profit_loss_rate_4 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_4(퇴직 연금)

irp_eval_amt_4 = irp_eval_amt_4.split("원")   ## 평가금액_4(퇴직 연금)
irp_eval_amt_4 = irp_eval_amt_4[0]
irp_eval_profit_loss_rate_4 = irp_eval_profit_loss_rate_4.split("%")   # 손익률(%)_4(퇴직 연금)
irp_eval_profit_loss_rate_4 = irp_eval_profit_loss_rate_4[0]  

self_house = ws_02_irp[f'L{rowNo}'].value  	        # 부동산(현아트빌 404) (05. 부동산)
rowNo = add_row + 11
insu_non_life_soo = ws_02_irp[f'L{rowNo}'].value    # 실비 보험(진수종) (06. 보험)
rowNo = add_row + 12
insu_non_life = ws_02_irp[f'L{rowNo}'].value  	    # 실비보험(06. 보험)
rowNo = add_row + 16
deposit_savings_sh = ws_02_irp[f'L{rowNo}'].value    # 01. 신한은행[예. 적금] (04. 은행 예. 적금)
rowNo = add_row + 17
deposit_savings_2 = ws_02_irp[f'L{rowNo}'].value     # 01. 은행 예. 적금_2(04. 은행 예. 적금)
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 10
irp_stock_nm_5 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_5(퇴직 연금)
irp_eval_amt_5 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_5(퇴직 연금)
irp_eval_profit_loss_5 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_5(퇴직 연금)
irp_eval_profit_loss_rate_5 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_5(퇴직 연금)

irp_eval_amt_5 = irp_eval_amt_5.split("원")   ## 평가금액_5(퇴직 연금)
irp_eval_amt_5 = irp_eval_amt_5[0]
irp_eval_profit_loss_rate_5 = irp_eval_profit_loss_rate_5.split("%")   # 손익률(%)_5(퇴직 연금)
irp_eval_profit_loss_rate_5 = irp_eval_profit_loss_rate_5[0]  
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 11
irp_stock_nm_6 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_6(퇴직 연금)
irp_eval_amt_6 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_6(퇴직 연금)
irp_eval_profit_loss_6 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_6(퇴직 연금)
irp_eval_profit_loss_rate_6 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_6(퇴직 연금)

irp_eval_amt_6 = irp_eval_amt_6.split("원")   ## 평가금액_6(퇴직 연금)
irp_eval_amt_6 = irp_eval_amt_6[0]
irp_eval_profit_loss_rate_6 = irp_eval_profit_loss_rate_6.split("%")   # 손익률(%)_6(퇴직 연금)
irp_eval_profit_loss_rate_6 = irp_eval_profit_loss_rate_6[0]  
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 12
irp_stock_nm_7 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_7(퇴직 연금)
irp_eval_amt_7 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_7(퇴직 연금)
irp_eval_profit_loss_7 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_7(퇴직 연금)
irp_eval_profit_loss_rate_7 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_7(퇴직 연금)

irp_eval_amt_7 = irp_eval_amt_7.split("원")   ## 평가금액_7(퇴직 연금)
irp_eval_amt_7 = irp_eval_amt_7[0]
irp_eval_profit_loss_rate_7 = irp_eval_profit_loss_rate_7.split("%")   # 손익률(%)_7(퇴직 연금)
irp_eval_profit_loss_rate_7 = irp_eval_profit_loss_rate_7[0]  
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 13
irp_stock_nm_8 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_8(퇴직 연금)
irp_eval_amt_8 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_8(퇴직 연금)
irp_eval_profit_loss_8 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_8(퇴직 연금)
irp_eval_profit_loss_rate_8 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_8(퇴직 연금)

irp_eval_amt_8 = irp_eval_amt_8.split("원")   ## 평가금액_8(퇴직 연금)
irp_eval_amt_8 = irp_eval_amt_8[0]
irp_eval_profit_loss_rate_8 = irp_eval_profit_loss_rate_8.split("%")   # 손익률(%)_8(퇴직 연금)
irp_eval_profit_loss_rate_8 = irp_eval_profit_loss_rate_8[0]  
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 14
irp_stock_nm_9 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_9(퇴직 연금)
irp_eval_amt_9 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_9(퇴직 연금)
irp_eval_profit_loss_9 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_9(퇴직 연금)
irp_eval_profit_loss_rate_9 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_9(퇴직 연금)

irp_eval_amt_9 = irp_eval_amt_9.split("원")   ## 평가금액_9(퇴직 연금)
irp_eval_amt_9 = irp_eval_amt_9[0]
irp_eval_profit_loss_rate_9 = irp_eval_profit_loss_rate_9.split("%")   # 손익률(%)_9(퇴직 연금)
irp_eval_profit_loss_rate_9 = irp_eval_profit_loss_rate_9[0]  
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 15
irp_stock_nm_10 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_10(퇴직 연금)
irp_eval_amt_10 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_10(퇴직 연금)
irp_eval_profit_loss_10 = ws_02_irp[f'E{rowNo}'].value  	    # 평가손익_10(퇴직 연금)
irp_eval_profit_loss_rate_10 = ws_02_irp[f'F{rowNo}'].value   # 손익률(%)_10(퇴직 연금)

irp_eval_amt_10 = irp_eval_amt_10.split("원")   ## 평가금액_10(퇴직 연금)
irp_eval_amt_10 = irp_eval_amt_10[0]
irp_eval_profit_loss_rate_10 = irp_eval_profit_loss_rate_10.split("%")   # 손익률(%)_10(퇴직 연금)
irp_eval_profit_loss_rate_10 = irp_eval_profit_loss_rate_10[0]  
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 16
irp_stock_nm_11 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_11(퇴직 연금)
irp_eval_amt_11 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_11(퇴직 연금)

irp_eval_profit_loss_11 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_11(퇴직 연금)
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_11] [평가금액_11]"+ str(irp_eval_amt_11) +"[평가손익_11]"+ str(type(irp_eval_profit_loss_11)) )

if irp_eval_profit_loss_11 != None  :   # 평가손익_11이 존재하면 
    irp_eval_profit_loss_11 = irp_eval_profit_loss_11
else: 
    irp_eval_profit_loss_11 = ''

irp_eval_profit_loss_rate_11 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_11퇴직 연금)
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_14] [손익률(%)_11]"+ str(type(irp_eval_profit_loss_rate_11)) )

if irp_eval_profit_loss_rate_11 != None  :   # 평가손익_11이 존재하면 
    irp_eval_profit_loss_rate_11 = irp_eval_profit_loss_rate_11
else: 
    irp_eval_profit_loss_rate_11 = ''
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_15] [손익률(%)_11]"+ str(irp_eval_profit_loss_rate_11) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 17
irp_stock_nm_12 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_12(퇴직 연금)
irp_eval_amt_12 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_12(퇴직 연금)

irp_eval_profit_loss_12 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_12(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_12] [평가금액_12]"+ str(irp_eval_amt_12) +"[평가손익_12]"+ str(type(irp_eval_profit_loss_12)) )

if irp_eval_profit_loss_12 != None  :   # 평가손익_12이 존재하면 
    irp_eval_profit_loss_12 = irp_eval_profit_loss_12
else: 
    irp_eval_profit_loss_12 = ''

irp_eval_profit_loss_rate_12 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_12퇴직 연금)

if irp_eval_profit_loss_rate_12 != None  :   # 평가손익_12이 존재하면 
    irp_eval_profit_loss_rate_12 = irp_eval_profit_loss_rate_12
else: 
    irp_eval_profit_loss_rate_12 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_15] [손익률(%)_12]"+ str(irp_eval_profit_loss_rate_12) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 18
irp_stock_nm_13 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_13(퇴직 연금)
irp_eval_amt_13 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_13(퇴직 연금)

irp_eval_profit_loss_13 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_13(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_13] [평가금액_13]"+ str(irp_eval_amt_13) +"[평가손익_13]"+ str(type(irp_eval_profit_loss_13)) )

if irp_eval_profit_loss_13 != None  :   # 평가손익_13이 존재하면 
    irp_eval_profit_loss_13 = irp_eval_profit_loss_13
else: 
    irp_eval_profit_loss_13 = ''

irp_eval_profit_loss_rate_13 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_13퇴직 연금)

if irp_eval_profit_loss_rate_13 != None  :   # 평가손익_13이 존재하면 
    irp_eval_profit_loss_rate_13 = irp_eval_profit_loss_rate_13
else: 
    irp_eval_profit_loss_rate_13 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_15] [손익률(%)_13]"+ str(irp_eval_profit_loss_rate_13) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 19
irp_stock_nm_14 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_14(퇴직 연금)
irp_eval_amt_14 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_14(퇴직 연금)

irp_eval_profit_loss_14 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_14(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_14] [평가금액_14]"+ str(irp_eval_amt_14) +"[평가손익_14]"+ str(type(irp_eval_profit_loss_14)) )

if irp_eval_profit_loss_14 != None  :   # 평가손익_14이 존재하면 
    irp_eval_profit_loss_14 = irp_eval_profit_loss_14
else: 
    irp_eval_profit_loss_14 = ''

irp_eval_profit_loss_rate_14 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_14퇴직 연금)

if irp_eval_profit_loss_rate_14 != None  :   # 평가손익_14이 존재하면 
    irp_eval_profit_loss_rate_14 = irp_eval_profit_loss_rate_14
else: 
    irp_eval_profit_loss_rate_14 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_15] [손익률(%)_14]"+ str(irp_eval_profit_loss_rate_14) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 20
irp_stock_nm_15 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_15(퇴직 연금)
irp_eval_amt_15 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_15(퇴직 연금)

irp_eval_profit_loss_15 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_15(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_15] [평가금액_15]"+ str(irp_eval_amt_15) +"[평가손익_15]"+ str(type(irp_eval_profit_loss_15)) )

if irp_eval_profit_loss_15 != None  :   # 평가손익_15이 존재하면 
    irp_eval_profit_loss_15 = irp_eval_profit_loss_15
else: 
    irp_eval_profit_loss_15 = ''

irp_eval_profit_loss_rate_15 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_15퇴직 연금)

if irp_eval_profit_loss_rate_15 != None  :   # 평가손익_15이 존재하면 
    irp_eval_profit_loss_rate_15 = irp_eval_profit_loss_rate_15
else: 
    irp_eval_profit_loss_rate_15 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_15] [손익률(%)_15]"+ str(irp_eval_profit_loss_rate_15) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 21
irp_stock_nm_16 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_16(퇴직 연금)
irp_eval_amt_16 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_16(퇴직 연금)

irp_eval_profit_loss_16 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_16(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_16] [평가금액_16]"+ str(irp_eval_amt_16) +"[평가손익_16]"+ str(type(irp_eval_profit_loss_16)) )

if irp_eval_profit_loss_16 != None  :   # 평가손익_16이 존재하면 
    irp_eval_profit_loss_16 = irp_eval_profit_loss_16
else: 
    irp_eval_profit_loss_16 = ''

irp_eval_profit_loss_rate_16 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_16퇴직 연금)

if irp_eval_profit_loss_rate_16 != None  :   # 평가손익_16이 존재하면 
    irp_eval_profit_loss_rate_16 = irp_eval_profit_loss_rate_16
else: 
    irp_eval_profit_loss_rate_16 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_17] [손익률(%)_16]"+ str(irp_eval_profit_loss_rate_16) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 22
irp_stock_nm_17 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_17(퇴직 연금)
irp_eval_amt_17 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_17(퇴직 연금)

irp_eval_profit_loss_17 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_17(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_17] [평가금액_17]"+ str(irp_eval_amt_17) +"[평가손익_17]"+ str(type(irp_eval_profit_loss_17)) )

if irp_eval_profit_loss_17 != None  :   # 평가손익_17이 존재하면 
    irp_eval_profit_loss_17 = irp_eval_profit_loss_17
else: 
    irp_eval_profit_loss_17 = ''

irp_eval_profit_loss_rate_17 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_17퇴직 연금)

if irp_eval_profit_loss_rate_17 != None  :   # 평가손익_17이 존재하면 
    irp_eval_profit_loss_rate_17 = irp_eval_profit_loss_rate_17
else: 
    irp_eval_profit_loss_rate_17 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_17] [손익률(%)_17]"+ str(irp_eval_profit_loss_rate_17) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 23
irp_stock_nm_18 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_18(퇴직 연금)
irp_eval_amt_18 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_18(퇴직 연금)

irp_eval_profit_loss_18 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_18(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_18] [평가금액_18]"+ str(irp_eval_amt_18) +"[평가손익_18]"+ str(type(irp_eval_profit_loss_18)) )

if irp_eval_profit_loss_18 != None  :   # 평가손익_18이 존재하면 
    irp_eval_profit_loss_18 = irp_eval_profit_loss_18
else: 
    irp_eval_profit_loss_18 = ''

irp_eval_profit_loss_rate_18 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_18퇴직 연금)

if irp_eval_profit_loss_rate_18 != None  :   # 평가손익_18이 존재하면 
    irp_eval_profit_loss_rate_18 = irp_eval_profit_loss_rate_18
else: 
    irp_eval_profit_loss_rate_18 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_19] [손익률(%)_18]"+ str(irp_eval_profit_loss_rate_18) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 24
irp_stock_nm_19 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_19(퇴직 연금)
irp_eval_amt_19 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_19(퇴직 연금)

irp_eval_profit_loss_19 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_19(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_19] [평가금액_19]"+ str(irp_eval_amt_19) +"[평가손익_19]"+ str(type(irp_eval_profit_loss_19)) )

if irp_eval_profit_loss_19 != None  :   # 평가손익_19이 존재하면 
    irp_eval_profit_loss_19 = irp_eval_profit_loss_19
else: 
    irp_eval_profit_loss_19 = ''

irp_eval_profit_loss_rate_19 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_19퇴직 연금)

if irp_eval_profit_loss_rate_19 != None  :   # 평가손익_19이 존재하면 
    irp_eval_profit_loss_rate_19 = irp_eval_profit_loss_rate_19
else: 
    irp_eval_profit_loss_rate_19 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_19] [손익률(%)_19]"+ str(irp_eval_profit_loss_rate_19) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 25
irp_stock_nm_20 = ws_02_irp[f'C{rowNo}'].value  	   # 상품명_20(퇴직 연금)
irp_eval_amt_20 = ws_02_irp[f'D{rowNo}'].value  	   # 평가금액_20(퇴직 연금)

irp_eval_profit_loss_20 = ws_02_irp[f'E{rowNo}'].value     # 평가손익_20(퇴직 연금)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_20] [평가금액_20]"+ str(irp_eval_amt_20) +"[평가손익_20]"+ str(type(irp_eval_profit_loss_20)) )

if irp_eval_profit_loss_20 != None  :   # 평가손익_20이 존재하면 
    irp_eval_profit_loss_20 = irp_eval_profit_loss_20
else: 
    irp_eval_profit_loss_20 = ''

irp_eval_profit_loss_rate_20 = ws_02_irp[f'F{rowNo}'].value     # 손익률(%)_20퇴직 연금)

if irp_eval_profit_loss_rate_20 != None  :   # 평가손익_20이 존재하면 
    irp_eval_profit_loss_rate_20 = irp_eval_profit_loss_rate_20
else: 
    irp_eval_profit_loss_rate_20 = ''
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_20] [손익률(%)_20]"+ str(irp_eval_profit_loss_rate_20) )
# ---------------------------------------------------------------------------------------------------------------------->


rowNo = add_row + 18
deposit_savings_3 = ws_02_irp[f'L{rowNo}'].value     # 01. 은행 예. 적금_3 (04. 은행 예. 적금) 
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_22] [01. 은행 예. 적금_2]"+ str(type(deposit_savings_2)) +"[01. 은행 예. 적금_3]"+ str(type(deposit_savings_3)) )
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_23] [01. 은행 예. 적금_2]"+ str(deposit_savings_2) +"[01. 은행 예. 적금_3]"+ str(deposit_savings_3) )

if deposit_savings_sh == None or deposit_savings_sh == '' :   # 02. 대여금이 널 이면
    deposit_savings_sh = 0
if int(deposit_savings_sh) > 0  :   # 02. 대여금이 존재하면
    deposit_savings_sh_nm = "01. 신한은행[예. 적금]"

if deposit_savings_2 == None or deposit_savings_2 == '' :   # 01. 은행 예. 적금_2이 널 이면
    deposit_savings_2 = 0

if deposit_savings_3 == None or deposit_savings_3 == '' :   # 01. 은행 예. 적금_3이 널 이면
    deposit_savings_3 = 0

deposit_savings = int(deposit_savings_sh) + int(deposit_savings_2) + int(deposit_savings_3)
deposit_savings_nm = "01. 은행 예. 적금([06])"
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_32] [01. 신한은행[예. 적금]]"+ str(deposit_savings_sh) +"[01. 은행 예. 적금_2]"+ str(deposit_savings_2) +"[01. 은행 예. 적금_3]"+ str(deposit_savings_3) )
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_41] [01. 은행 예. 적금]"+ str(deposit_savings) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 24
rental_fee = ws_02_irp[f'L{rowNo}'].value   # 02. 대여금(문태용[24.05까지])
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_12] [rental_fee]"+ str(rental_fee) )

if rental_fee == None or rental_fee == '' :   # 02. 대여금이 널 이면
    rental_fee = 0
if int(rental_fee) > 0 :   # 02. 대여금이 존재하면
    rental_fee_nm = "02. 대여금(문태용[24.05까지)"

tot_deposit_savings = int(deposit_savings) + int(rental_fee)  # 1. 총 예적금
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_21_51] [1. 총 예적금]"+ str(tot_deposit_savings) )
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 29
mon_expendit = ws_02_irp[f'L{rowNo}'].value   # 91. 월 지출(A)
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 30
mon_income = ws_02_irp[f'L{rowNo}'].value   # 92. 월 수입(B)
# ---------------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 33
settlement = ws_02_irp[f'L{rowNo}'].value   # 결산
# ===============================================================================================================================
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


print("\n\n [@_T] ■■■ [/ast_vrfc.py] ==> [T_01] ■■■■■■ [######################### [22. 채권 Tab(기준 정보) 작업 Start] #########################] ■■■■■■ ")

ws_03_bond = wb["채권"]   # "채권" Sheet(Tab)에 접근 @@@@@@@@ ===========>
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_22_01] [채권 Tab Sheet(Tab)에 접근 @@@@@@@@ ===========>]")

# 채권 보유현황 @@@ 
add_row = 0     # 줄 추가
rowNo = add_row + 6
prchs_amt78 = ws_03_bond[f'E{rowNo}'].value        # 04. 매입금액(78. 발행어음 만기형_개인(181~270))
eval_amt78 = ws_03_bond[f'F{rowNo}'].value         # 05. 평가금액(78. 발행어음 만기형_개인(181~270))
eval_profit_loss78 = ws_01_actBalnc[f'G{rowNo}'].value   # 06. 평가손익(78. 발행어음 만기형_개인(181~270)) 
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_31] [04. 매입금액(79. 발행어음 CMA_개인)_00]" )

rowNo = add_row + 7
cmaRowNo = add_row + 7
if ws_03_bond[f'E{rowNo}'].value != None :    # 04. 매입금액(79. 발행어음 CMA_개인)이 널이 아니면
    cma_prchs_amt79 = ws_03_bond[f'E{rowNo}'].value   # 04. 매입금액(79. 발행어음 CMA_개인)  
else:
    cma_prchs_amt79 = 0   # 04. 매입금액(79. 발행어음 CMA_개인)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_32] [04. 매입금액(79. 발행어음 CMA_개인)]"+ str(ws_03_bond[f'E{rowNo}'].value) +"[04. 매입금액(79. 발행어음 CMA_개인)_22]"+ str(cma_prchs_amt79) )

if ws_03_bond[f'F{rowNo}'].value != None :    # 05. 평가금액(79. 발행어음 CMA_개인)이 널이 아니면
    cma_eval_amt79 = ws_03_bond[f'F{rowNo}'].value   # 05. 평가금액(79. 발행어음 CMA_개인)  
else:
    cma_eval_amt79 = 0   # 05. 평가금액(79. 발행어음 CMA_개인)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_33] [05. 평가금액(79. 발행어음 CMA_개인)]"+ str(cma_eval_amt79) )

eval_profit_loss79 = ws_01_actBalnc[f'G{rowNo}'].value   # 06. 평가손익(79. 발행어음 CMA_개인) 
# ---------------------------------------------------------------------------------------------------------------------->
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


print("\n\n [@_T] ■■■ [/ast_vrfc.py] ==> [T_01] ■■■■■■ [######################### [02. 총자산(24Y) Tab 작업 Start] #########################] ■■■■■■ ")

totAstYear = "02. 총자산("+ str(astYYM[2:4]) +"Y)"
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_01] [자산(202*) Sheet(Tab)에 접근 @@@@@@@@ ===========>]")
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_01] [00. totAstYear]"+ str(totAstYear) +"[자산 년]"+ str(input_astYYYYMM[0]) +"[자산 월]"+ str(input_astYYYYMM[1]) )

if int(input_astYYYYMM[1]) > 11  :   # 자산 월이 12월 이면
    wsTot = wb.create_sheet(totAstYear, 0)   # 엑셀 Sheet 생성
    wsTot = wb[totAstYear]    # "자산(2023)" Sheet(Tab)에 접근 @@@@@@@@ ===========>
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_02] [엑셀 Sheet 생성] @@@@@@@@ ===========>]"+ str(totAstYear) )
else :
    wsTot = wb[totAstYear]    # "02. 총자산(24Y)" Sheet(Tab)에 접근 @@@@@@@@ ===========>
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_03] [엑셀 Sheet(Tab)에 접근] @@@@@@@@ ===========>]"+ str(totAstYear) )
#print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_04] [엑셀 Sheet 명] @@@@@@@@ ===========>]"+ str(wb.sheetnames) )

wsTot.column_dimensions["B"].width = 15  # A열의 너비를 50로 설정
wsTot.column_dimensions["L"].width = 15  # A열의 너비를 10로 설정
wsTot.column_dimensions["N"].width = 50  # G열의 너비를 18로 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_05_1] [prev_B7_amt]"+ str(prev_B7_amt) +"[prev_B14_amt]"+ str(prev_B14_amt) )
# ---------------------------------------------------------------------------------------------------------------------->

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))


# 02. 총자산(24Y) Tab 줄 추가, 셀 병합
insertRow = 25 + 2      # 월별 자산 추가 줄(17줄 + 2줄[상위 빈 줄])
wsTot.insert_rows(3, insertRow)    # 월별 자산 제목 아래 부터 줄부터 18줄 추가(3번째 줄 위치에 19줄을 추가 ==> 17줄)
print("\n\n[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_11] [02. 총자산(24Y) Tab 줄 추가, 셀 병합 처리]===========>" )

wsTot.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)     # '02. 총자산(24Y)' 셀 병합("A1:I1") 
wsTot.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)     # '2024.06 자산' 셀 병합("A3:H3")   ==> 3
wsTot.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)     # '01. 주식 월별 보유 현황' 셀 병합("A4:I4") 
wsTot.merge_cells(start_row=4, start_column=11, end_row=4, end_column=18)   # '03. IRP 월별 보유 현황(미래에셋 증권 IRP)("K4:R4")
wsTot.merge_cells(start_row=8, start_column=11, end_row=8, end_column=17)   # '03-1. 퇴직신탁 보유현황("K8:Q8") 
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_50] [item_0번째] [01. start_row]3")

addRow = 0
addRowFrst = int(insertRow) + 3         # 줄 추가_01[2024.06 자산 칼럼] = 19 + 3줄[하위 빈 줄])
addRowFrst2 = int(insertRow) + 25 + 2   # 줄 추가_02[결산 칼럼] = 19 + 19줄[하위 빈 줄])

for i in range(1, 13):
    addRow = int(addRow) + int(addRowFrst)     # 줄 추가_03 = 22 + 3 = 41
    if i > 1 : addRow = int(addRow) - 3 
    wsTot.merge_cells(start_row=addRow, start_column=1, end_row=addRow, end_column=8)         # '2024.06 자산' 셀 병합("A3:H3")    ==> 30 ~
    wsTot.merge_cells(start_row=addRow+1, start_column=1, end_row=addRow+1, end_column=9)     # '01. 주식 월별 보유 현황' 셀 병합("A4:I4") 
    wsTot.merge_cells(start_row=addRow+1, start_column=11, end_row=addRow+1, end_column=18)   # '03. IRP 월별 보유 현황(미래에셋 증권 IRP)("K4:R4") 
    wsTot.merge_cells(start_row=addRow+5, start_column=11, end_row=addRow+5, end_column=17)   # '03-1. 퇴직신탁 보유현황("K8:Q8")
    wsTot.merge_cells(start_row=addRow+6, start_column=1, end_row=addRow+6, end_column=2)     # '합계("A9:B9")
    print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_51] [item_번째]"+ str(i) +"[01. start_row]"+ str(addRow) +"[02. addRow+1]"+ str(addRow+1)  +"[05. addRow+5]"+ str(addRow+5) )
# -------------------------------------------------------------------------------------------->

# if int(input_astYYYYMM[1]) > 11  :   # 자산 월이 12월 이면
# wsTot.merge_cells("A1:I1")   # '02. 총자산(24Y)' 셀 병합("A1:I1") 
wsTot[f'A{rowNo}'].value = "02. 총자산("+ str(astYYM[0:4]) +")"  # 01. 필드에 값 쓰기 ---> 02. 총자산 ■
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=13)   # 02. 필드 글자 설정 
wsTot[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
wsTot[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
wsTot[f'A{rowNo}'].fill = blueDark2Fill    # 05. 필드 배경색 설정
# ------------------------------------------------------------------------------------------------------------->

add_row = 0     # 줄 추가
rowNo = add_row + 3     # "2023.12 자산" 칼럼 Row(02. 총자산(24Y) Tab에서)
# wsTot.merge_cells("A3:H3")   # A2부터 H3까지 셀을 싱글 셀로 병합
wsTot[f'A{rowNo}'].value = astYYM +" 자산"   # 01. 필드에 값 쓰기 ---> 자산 년월 ■
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=12)   # 02. 필드 글자 설정 
wsTot[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'A{rowNo}'].fill = orangeFill      # 05. 필드 배경색 설정

wsTot[f'I{rowNo}'].value = lastLastDt     # 01. 필드에 값 쓰기 ---> 024.01.01(작성 년월일) ■
wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9)     # 02. 필드 글자 설정 
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')    # 03. 필드 정렬 설정
wsTot[f'I{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'I{rowNo}'].fill = orangeFill      # 05. 필드 배경색 설정 
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 4    # "01. 주식 월별 보유 현황" 칼럼 Row(02. 총자산(24Y) Tab에서)
# wsTot.merge_cells("A4:I4")   # A1부터 I1까지 셀을 싱글 셀로 병합
wsTot[f'A{rowNo}'].value = "01. 주식 월별 보유 현황"   # 01. 필드에 값 쓰기 ---> 01. 주식 월별 보유 현황 ■
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=9)   # 02. 필드 글자 설정 
wsTot[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'I{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'A{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정

wsTot[f'K{rowNo}'].value = "03. IRP 월별 보유 현황(미래에셋 증권 IRP)"   # 01. 필드에 값 쓰기 ---> 01. 주식 월별 보유 현황 ■
# wsTot.merge_cells("K4:R4")   # A1부터 I1까지 셀을 싱글 셀로 병합
wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=True, size=9)   # 02. 필드 글자 설정 
wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
wsTot[f'K{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정
wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'R{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 5    # "01. 주식 월별 보유 현황 Head" 칼럼 Row(02. 총자산(24Y) Tab에서) 
wsTot[f'A{rowNo}'].value = "주식코드"   # 01. 필드에 값 쓰기 ---> 주식코드 ■
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
wsTot[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'A{rowNo}'].fill = grayFill      # 05. 필드 배경색 설정

wsTot[f'B{rowNo}'].value = '상품명'          # 01. 필드에 값 쓰기 ---> 상품명 ■
wsTot[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'B{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'B{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'C{rowNo}'].value = '보유수량'          # 01. 필드에 값 쓰기 ---> 보유수량 ■
wsTot[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'C{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'C{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'D{rowNo}'].value = '현재가'          # 01. 필드에 값 쓰기 ---> 현재가 ■
wsTot[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'D{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'E{rowNo}'].value = '평균매입가'     # 01. 필드에 값 쓰기 ---> 평균매입가 ■
wsTot[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'E{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'E{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'F{rowNo}'].value = '매입금액'          # 01. 필드에 값 쓰기 ---> 매입금액 ■
wsTot[f'F{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'F{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'F{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'G{rowNo}'].value = '평가금액'          # 01. 필드에 값 쓰기 ---> 평가금액 ■
wsTot[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'G{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'G{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'H{rowNo}'].value = '평가손익'     # 01. 필드에 값 쓰기 ---> 평가손익 ■
wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'H{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'H{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'I{rowNo}'].value = '손익률(%)'     # 01. 필드에 값 쓰기 ---> 손익률(%) ■
wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'I{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'I{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

wsTot[f'K{rowNo}'].value = "No."   # 01. 필드에 값 쓰기 ---> No. ■ 
wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'K{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 
wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'L{rowNo}'].value = "계좌번호"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'L{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정

wsTot[f'M{rowNo}'].value = "제도"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'M{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정
wsTot[f'N{rowNo}'].value = "상품명"   # 01. 필드에 값 쓰기 ---> 상품명 ■ 
wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'N{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'N{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 

wsTot[f'O{rowNo}'].value = "매입금액"   # 01. 필드에 값 쓰기 ---> 매입금액 ■ 
wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'O{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 

wsTot[f'P{rowNo}'].value = "평가금액"   # 01. 필드에 값 쓰기 ---> 평가금액 ■ 
wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'P{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 

wsTot[f'Q{rowNo}'].value = "평가손익"   # 01. 필드에 값 쓰기 ---> 평가손익 ■ 
wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'Q{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 

wsTot[f'R{rowNo}'].value = "손익률(%)"   # 01. 필드에 값 쓰기 ---> 손익률(%) ■ 
wsTot[f'R{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'R{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'R{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'R{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 6    # "01. 주식 월별 보유 현황 Body_1" 칼럼 Row(02. 총자산(24Y) Tab에서) 
wsTot[f'A{rowNo}'].value = "005380"   # 01. 필드에 값 쓰기 ---> 주식코드 Value ■ --> 주식(현대차)
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
# wsTot[f'A{rowNo}'].fill = violetFill      # 05. 필드 배경색 설정 ===> TEST @@@

wsTot[f'B{rowNo}'].value = '현대차'          # 01. 필드에 값 쓰기 ---> 상품명 Value ■
wsTot[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정 
wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'C{rowNo}'].value = format(posses_qty1, ',')        # 01. 필드에 값 쓰기 ---> 보유수량 Value ■ --> 주식(현대차)
wsTot[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'D{rowNo}'].value = format(cur_amt1, ',')          # 01. 필드에 값 쓰기 ---> 현재가 Value ■ --> 주식(현대차)
wsTot[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].fill = yellowFill      # 05. 필드 배경색 설정

wsTot[f'E{rowNo}'].value = format(ave_prchs_amt1, ',')    # 01. 필드에 값 쓰기 ---> 평균매입가 Value ■
wsTot[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'F{rowNo}'].value = format(prchs_amt1, ',')        # 01. 필드에 값 쓰기 ---> 매입금액 Value ■
wsTot[f'F{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'F{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'G{rowNo}'].value = format(eval_amt1, ',')         # 01. 필드에 값 쓰기 ---> 평가금액 Value ■
wsTot[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'G{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'H{rowNo}'].value = format(eval_profit_loss1, ',')      # 01. 필드에 값 쓰기 ---> 평가손익 Value ■
wsTot[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_05] [(현대차)_손익률(%)]"+ str(eval_profit_loss_rate1) )

wsTot[f'I{rowNo}'].value = format(eval_profit_loss_rate1, ',')    # 01. 필드에 값 쓰기 ---> 손익률(%) Value ■ --> 주식(현대차)
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'I{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

if eval_amt1 > 0  :   # 평가금액이 양이면
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    
wsTot[f'K{rowNo}'].value = "1"   # 01. 필드에 값 쓰기 ---> No. ■ 
wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

wsTot[f'N{rowNo}'].value = "퇴직연금"   # 01. 필드에 값 쓰기 ---> 상품명 ■ 
wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

wsTot[f'O{rowNo}'].value = format(irp_prchs_amt, ',')      # 01. 필드에 값 쓰기 ---> 매입금액(퇴직 연금) Value ■
wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

wsTot[f'P{rowNo}'].value = format(irp_eval_amt, ',')      # 01. 필드에 값 쓰기 ---> 평가금액(퇴직 연금) Value ■
wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

wsTot[f'Q{rowNo}'].value = format(irp_eval_profit_loss, ',')      # 01. 필드에 값 쓰기 ---> 평가손익(퇴직 연금) Value ■
wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=True, size=9)   # 02. 필드 글자 설정  
wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

wsTot[f'R{rowNo}'].value = format(irp_eval_profit_loss_rate * 100, ',')   # 01. 필드에 값 쓰기 ---> 손익률(%) (퇴직 연금) Value ■
wsTot[f'R{rowNo}'].font = Font(name="돋움", bold=True, size=9)   # 02. 필드 글자 설정  
wsTot[f'R{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'R{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

if irp_eval_profit_loss > 0  :   # 평가손익이 양이면
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    wsTot[f'R{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    wsTot[f'R{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 7    # "01. 주식 월별 보유 현황 Body_2" 칼럼 Row(02. 총자산(24Y) Tab에서) 
wsTot[f'A{rowNo}'].value = "068270"   # 01. 필드에 값 쓰기 ---> 주식코드 Value ■ --> 주식(셀트리온)
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'A{rowNo}'].alignment = Alignment(horizontal='left', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'A{rowNo}'].fill = whiteFill      # 05. 필드 배경색(흰색) 설정 ===> TEST @@@

wsTot[f'B{rowNo}'].value = '셀트리온'          # 01. 필드에 값 쓰기 ---> 상품명 Value ■
wsTot[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정 
wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'C{rowNo}'].value = format(posses_qty2, ',')        # 01. 필드에 값 쓰기 ---> 보유수량 Value ■ --> 주식(셀트리온)
wsTot[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'D{rowNo}'].value = format(cur_amt2, ',')          # 01. 필드에 값 쓰기 ---> 현재가 Value ■ --> 주식(셀트리온)
wsTot[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].fill = yellowFill        # 05. 필드 배경색 설정

wsTot[f'E{rowNo}'].value = format(ave_prchs_amt2, ',')    # 01. 필드에 값 쓰기 ---> 평균매입가 Value ■
wsTot[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'F{rowNo}'].value = format(prchs_amt2, ',')        # 01. 필드에 값 쓰기 ---> 매입금액 Value ■
wsTot[f'F{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'F{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'G{rowNo}'].value = format(eval_amt2, ',')         # 01. 필드에 값 쓰기 ---> 평가금액 Value ■
wsTot[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'G{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'H{rowNo}'].value = format(eval_profit_loss2, ',')      # 01. 필드에 값 쓰기 ---> 평가손익 Value ■
wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_06] [(셀트리온)_손익률(%)]"+ str(eval_profit_loss_rate2) )

wsTot[f'I{rowNo}'].value = format(eval_profit_loss_rate2, ',')    # 01. 필드에 값 쓰기 ---> 손익률(%) Value ■ --> 주식(셀트리온)
wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'I{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_11] [01. 현대차_손익률]"+ str(eval_profit_loss_rate1) +"[02. 셀트리온_손익]"+ str(eval_profit_loss_rate2) )

if eval_profit_loss2 > 0  :   # 평가손익이 양이면
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색) 
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 8    # "01. 주식 월별 보유 현황 Body_3" 칼럼 Row(02. 총자산(24Y) Tab에서) 
wsTot[f'A{rowNo}'].value = "096770"   # 01. 필드에 값 쓰기 ---> 주식코드 Value ■ --> 주식(SK이노베이션)
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'B{rowNo}'].value = 'SK이노베이션'      # 01. 필드에 값 쓰기 ---> 상품명 Value ■
wsTot[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정 
wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'C{rowNo}'].value = format(posses_qty3, ',')        # 01. 필드에 값 쓰기 ---> 보유수량 Value ■ --> 주식(SK이노베이션)
wsTot[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'D{rowNo}'].value = format(cur_amt3, ',')          # 01. 필드에 값 쓰기 ---> 현재가 Value ■ --> 주식(SK이노베이션)
wsTot[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].fill = yellowFill        # 05. 필드 배경색 설정

wsTot[f'E{rowNo}'].value = format(ave_prchs_amt3, ',')    # 01. 필드에 값 쓰기 ---> 평균매입가 Value ■
wsTot[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'F{rowNo}'].value = format(prchs_amt3, ',')        # 01. 필드에 값 쓰기 ---> 매입금액 Value ■
wsTot[f'F{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'F{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'G{rowNo}'].value = format(eval_amt3, ',')         # 01. 필드에 값 쓰기 ---> 평가금액 Value ■
wsTot[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'G{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'H{rowNo}'].value = format(eval_profit_loss3, ',')      # 01. 필드에 값 쓰기 ---> 평가손익 Value ■
wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_07] [(SK이노베이션)_손익률(%)]"+ str(eval_profit_loss_rate3) )

wsTot[f'I{rowNo}'].value = format(eval_profit_loss_rate3, ',')    # 01. 필드에 값 쓰기 ---> 손익률(%) Value ■ --> 주식(SK이노베이션)
# eval_profit_loss_rate3 = eval_profit_loss_rate3 * 10    # SK이노베이션)_손익률(%)
# # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_09] [(SK이노베이션)_손익률(%)]"+  f'{eval_profit_loss_rate3:.2f}' )
# wsTot[f'I{rowNo}'].value = f'{eval_profit_loss_rate3:.2f}'   # 01. 필드에 값 쓰기 ---> 손익률(%) Value ■ --> 주식(SK이노베이션)
wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'I{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

if eval_profit_loss3 > 0  :   # 평가손익이 양이면
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색) 

wsTot[f'K{rowNo}'].value = "03-1. 퇴직신탁 보유현황"   # 01. 필드에 값 쓰기 ---> 03-1. 퇴직신탁 보유현황) ■
# wsTot.merge_cells("K8:Q8")   # A1부터 I1까지 셀을 싱글 셀로 병합
wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=True, size=9)   # 02. 필드 글자 설정 
wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
wsTot[f'K{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정
wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 9    # "01. 주식 월별 보유 현황 Body_4" 칼럼 Row(02. 총자산(24Y) Tab에서)
# wsTot.merge_cells("A9:B9")   # A1부터 I1까지 셀을 싱글 셀로 병합
wsTot[f'A{rowNo}'].value = "합계"   # 01. 필드에 값 쓰기 ---> 합계 ■ --> 합계
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'A{rowNo}'].fill = orangeWeekFill        # 05. 필드 배경색 설정

wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'B{rowNo}'].fill = orangeWeekFill        # 05. 필드 배경색 설정

posses_qty = posses_qty1 + posses_qty2 + posses_qty3
wsTot[f'C{rowNo}'].value = format(posses_qty, ',')        # 01. 필드에 값 쓰기 ---> 보유수량 Value ■ 
wsTot[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
wsTot[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'C{rowNo}'].fill = orangeWeekFill        # 05. 필드 배경색 설정

cur_amt = cur_amt1 + cur_amt2 + cur_amt3
wsTot[f'D{rowNo}'].value = format(cur_amt, ',')          # 01. 필드에 값 쓰기 ---> 현재가 Value ■ 
wsTot[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
wsTot[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].fill = orangeWeekFill        # 05. 필드 배경색 설정

ave_prchs_amt = ave_prchs_amt1 + ave_prchs_amt2 + ave_prchs_amt3
wsTot[f'E{rowNo}'].value = format(ave_prchs_amt, ',')    # 01. 필드에 값 쓰기 ---> 평균매입가 Value ■
wsTot[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'E{rowNo}'].fill = orangeWeekFill        # 05. 필드 배경색 설정

prchs_amt = prchs_amt1 + prchs_amt2 + prchs_amt3
wsTot[f'F{rowNo}'].value = format(prchs_amt, ',')        # 01. 필드에 값 쓰기 ---> 매입금액 Value ■
wsTot[f'F{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'F{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'F{rowNo}'].fill = orangeWeekFill        # 05. 필드 배경색 설정

eval_amt = eval_amt1 + eval_amt2 + eval_amt3
wsTot[f'G{rowNo}'].value = format(eval_amt, ',')         # 01. 필드에 값 쓰기 ---> 평가금액 Value ■
wsTot[f'G{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")       # 02. 필드 글자 설정
wsTot[f'G{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'G{rowNo}'].fill = orangeWeekFill        # 05. 필드 배경색 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_10] [평가금액]"+ str(eval_amt) )

# 평가손익(원) = 평가금액 - 매입금액
eval_profit_loss =  eval_amt - prchs_amt
wsTot[f'H{rowNo}'].value = format(eval_profit_loss, ',')      # 01. 필드에 값 쓰기 ---> 평가손익 Value ■
wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'H{rowNo}'].border = thin_border     # 04. 필드 테두리 설정
wsTot[f'H{rowNo}'].fill = orangeWeekFill    # 05. 필드 배경색 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_11] [평가손익(원)]"+ str(eval_profit_loss) )

# 손익률(%)= (평가금액 - 매입금액) / 매입금액
eval_profit_loss_rate =  (eval_profit_loss / prchs_amt) * 100
wsTot[f'I{rowNo}'].value = f'{eval_profit_loss_rate:.2f}'   # 01. 필드에 값 쓰기 ---> 손익률(%) Value ■ --> 합계
wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'I{rowNo}'].border = thin_border     # 04. 필드 테두리 설정
wsTot[f'I{rowNo}'].fill = orangeWeekFill    # 05. 필드 배경색 설정

if eval_profit_loss > 0  :   # 평가손익이 양이면
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)

wsTot[f'K{rowNo}'].value = "No."   # 01. 필드에 값 쓰기 ---> No. ■ 
wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'K{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 
wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'L{rowNo}'].value = "계좌번호"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'L{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정

wsTot[f'M{rowNo}'].value = "제도"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'M{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정

wsTot[f'N{rowNo}'].value = "상품명"   # 01. 필드에 값 쓰기 ---> 상품명 ■ 
wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'N{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'N{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 

wsTot[f'O{rowNo}'].value = "평가금액"   # 01. 필드에 값 쓰기 ---> 평가금액 ■ 
wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'O{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 

wsTot[f'P{rowNo}'].value = "평가손익"   # 01. 필드에 값 쓰기 ---> 평가손익 ■ 
wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
wsTot[f'P{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 

wsTot[f'Q{rowNo}'].value = "손익률(%)"   # 01. 필드에 값 쓰기 ---> 손익률(%) ■ 
wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'Q{rowNo}'].fill = grayFill   # 05. 필드 배경색 설정 
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 10		# "01. 주식 월별 보유 현황 Body_4" 칼럼 Row(02. 총자산(24Y) Tab에서)
# wsTot.merge_cells("A10:C10")   # A10부터 C10까지 셀을 싱글 셀로 병합
wsTot[f'A{rowNo}'].value = "02. 발행어음 만기형_개인(181~270)"   # 01. 필드에 값 쓰기 ---> 합계 ■ --> 02. 발행어음 
wsTot[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
wsTot[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'F{rowNo}'].value = format(prchs_amt78, ',')        # 01. 필드에 값 쓰기 ---> 매입금액 Value ■  --> 02. 발행어음 
wsTot[f'F{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'F{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'F{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

wsTot[f'G{rowNo}'].value = format(eval_amt78, ',')         # 01. 필드에 값 쓰기 ---> 평가금액 Value ■
wsTot[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
wsTot[f'G{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_13] [평가금액]"+ str(eval_amt) )

# 평가손익(원) = 평가금액 - 매입금액
eval_profit_loss78 =  eval_amt78 - prchs_amt78
wsTot[f'H{rowNo}'].value = format(eval_profit_loss78, ',')      # 01. 필드에 값 쓰기 ---> 평가손익 Value ■
wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_14] [평가손익(원)]"+ str(eval_profit_loss) )

# 손익률(%)= (평가금액 - 매입금액) / 매입금액
eval_profit_loss_rate78 =  (eval_profit_loss78 / prchs_amt78) * 100
wsTot[f'I{rowNo}'].value = f'{eval_profit_loss_rate78:.2f}'   # 01. 필드에 값 쓰기 ---> 손익률(%) Value ■ --> 02. 발행어음 
wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
wsTot[f'I{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
wsTot[f'I{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

if eval_profit_loss > 0  :   # 평가손익이 양이면
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    wsTot[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    wsTot[f'I{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)

if irp_stock_nm_1 != None and irp_stock_nm_1 != '' :   # 상품명_2(퇴직 연금)이 존재하면 
    wsTot[f'K{rowNo}'].value = "1"   # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_1       # 01. 필드에 값 쓰기 ---> 상품명_1(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_1       # 01. 필드에 값 쓰기 ---> 평가금액_1(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_1, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_1(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_15] [손익률(%)_1(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_1      # 01. 필드에 값 쓰기 ---> 손익률(%)_1(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_1 > 0  :   # 평평가손익_1이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_2 != None and irp_stock_nm_2 != '' :   # 상품명_2(퇴직 연금)이 존재하면 
    rowNo = add_row + 11		# "03-1. 퇴직신탁 보유현황 Body_4" 칼럼 Row(02. 총자산(24Y) Tab에서)     ---> 상품명_2(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "2"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_2       # 01. 필드에 값 쓰기 ---> 상품명_2(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_2       # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_2, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_16] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_2      # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_2 > 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_3 != None and irp_stock_nm_3 != '' :   # 상품명_3(퇴직 연금)이 존재하면 
    rowNo = add_row + 12	# "03-1. 퇴직신탁 보유현황 칼럼 Row(02. 총자산(24Y) Tab에서)     ---> 상품명_3(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "3"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_3      # 01. 필드에 값 쓰기 ---> 상품명_2(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_3      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_3, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_15] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_3     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_3 > 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_4 != None and irp_stock_nm_4 != '' :   # 상품명_4(퇴직 연금)이 존재하면 
    rowNo = add_row + 13		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)      ---> 상품명_4(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "4"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_4      # 01. 필드에 값 쓰기 ---> 상품명_2(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_4      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_4, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_17] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_4     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_4 > 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_5 != None and irp_stock_nm_5 != '' :   # 상품명_5(퇴직 연금)이 존재하면 
    rowNo = add_row + 14		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)     ---> 상품명_5(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "5"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_5      # 01. 필드에 값 쓰기 ---> 상품명_5(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_5      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_5, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_18] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_5     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_5 > 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_6 != None and irp_stock_nm_6 != '' :   # 상품명_6(퇴직 연금)이 존재하면 
    rowNo = add_row + 15		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)      ---> 상품명_6(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "6"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_6      # 01. 필드에 값 쓰기 ---> 상품명_2(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_6      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_6, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_19] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_6     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_6 > 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_7 != None and irp_stock_nm_7 != '' :   # 상품명_7(퇴직 연금)이 존재하면
    rowNo = add_row + 16		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)    ---> 상품명_7(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "7"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_7      # 01. 필드에 값 쓰기 ---> 상품명_7(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_7      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_7, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_19] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_7     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_7 > 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_8 != None and irp_stock_nm_8 != '' :   # 상품명_8(퇴직 연금)이 존재하면
    rowNo = add_row + 17		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)     ---> 상품명_8(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "8"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_8      # 01. 필드에 값 쓰기 ---> 상품명_8(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_8      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_8, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_20] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_8     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_8> 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_9 != None and irp_stock_nm_9 != '' :   # 상품명_10(퇴직 연금)이 존재하면
    rowNo = add_row + 18		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_09(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "9"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_9      # 01. 필드에 값 쓰기 ---> 상품명_2(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_9      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_9, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_21] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_9     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_9> 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_10 != None and irp_stock_nm_10 != '' :   # 상품명_10(퇴직 연금)이 존재하면
    rowNo = add_row + 19		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)    ---> 상품명_10(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "10"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_10      # 01. 필드에 값 쓰기 ---> 상품명_2(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_10      # 01. 필드에 값 쓰기 ---> 평가금액_2(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_10, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_2(퇴직 연금) Value ■
    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_10     # 01. 필드에 값 쓰기 ---> 손익률(%)_2(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_10> 0  :   # 평평가손익_2이 양이면
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_11 != None and irp_stock_nm_11 != '' :   # 상품명_11(퇴직 연금)이 존재하면
    rowNo = add_row + 20		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)     ---> 상품명_11(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "11"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_11      # 01. 필드에 값 쓰기 ---> 상품명_11(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_11      # 01. 필드에 값 쓰기 ---> 평가금액_11(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_11) )

    if irp_eval_profit_loss_11 != ''  :   # 평가손익_11이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_11, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_11(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_11     # 01. 필드에 값 쓰기 ---> 평가손익_11(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_11     # 01. 필드에 값 쓰기 ---> 손익률(%)_11(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_11 == ''  :   # 평가손익_11이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_11 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [상품명_12(퇴직 연금)]"+ str(irp_stock_nm_12) )

if irp_stock_nm_12 != None and irp_stock_nm_12 != '' :   # 상품명_12(퇴직 연금)이 존재하면
    rowNo = add_row + 21		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)  ---> 상품명_12(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "12"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_12      # 01. 필드에 값 쓰기 ---> 상품명_12(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_12      # 01. 필드에 값 쓰기 ---> 평가금액_12(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_12) )

    if irp_eval_profit_loss_12 != ''  :   # 평가손익_12이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_12, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_12(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_12     # 01. 필드에 값 쓰기 ---> 평가손익_12(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_12     # 01. 필드에 값 쓰기 ---> 손익률(%)_12(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_12 == ''  :   # 평가손익_12이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_12 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_13 != None and irp_stock_nm_13 != '' :   # 상품명_13(퇴직 연금)이 존재하면
    rowNo = add_row + 22		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_13(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "13"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_13      # 01. 필드에 값 쓰기 ---> 상품명_13(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_13      # 01. 필드에 값 쓰기 ---> 평가금액_13(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_13) )

    if irp_eval_profit_loss_13 != ''  :   # 평가손익_13이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_13, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_13(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_13     # 01. 필드에 값 쓰기 ---> 평가손익_13(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_13     # 01. 필드에 값 쓰기 ---> 손익률(%)_13(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_13 == ''  :   # 평가손익_13이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_13 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_14 != None and irp_stock_nm_14 != '' :   # 상품명_14(퇴직 연금)이 존재하면
    rowNo = add_row + 23		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_14(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "14"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_14      # 01. 필드에 값 쓰기 ---> 상품명_14(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_14      # 01. 필드에 값 쓰기 ---> 평가금액_14(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_14) )

    if irp_eval_profit_loss_14 != ''  :   # 평가손익_14이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_14, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_14(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_14     # 01. 필드에 값 쓰기 ---> 평가손익_14(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_14     # 01. 필드에 값 쓰기 ---> 손익률(%)_14(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_14 == ''  :   # 평가손익_14이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_14 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_15 != None and irp_stock_nm_15 != '' :   # 상품명_15(퇴직 연금)이 존재하면
    rowNo = add_row + 24		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_15(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "15"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_15      # 01. 필드에 값 쓰기 ---> 상품명_15(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_15      # 01. 필드에 값 쓰기 ---> 평가금액_15(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_15) )

    if irp_eval_profit_loss_15 != ''  :   # 평가손익_15이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_15, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_15(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_15     # 01. 필드에 값 쓰기 ---> 평가손익_15(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_15     # 01. 필드에 값 쓰기 ---> 손익률(%)_15(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_15 == ''  :   # 평가손익_15이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_15 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_16 != None and irp_stock_nm_16 != '' :   # 상품명_16(퇴직 연금)이 존재하면
    rowNo = add_row + 25		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_16(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "16"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_16      # 01. 필드에 값 쓰기 ---> 상품명_16(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_16      # 01. 필드에 값 쓰기 ---> 평가금액_16(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_16) )

    if irp_eval_profit_loss_16 != ''  :   # 평가손익_16이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_16, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_16(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_16     # 01. 필드에 값 쓰기 ---> 평가손익_16(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_16     # 01. 필드에 값 쓰기 ---> 손익률(%)_16(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_16 == ''  :   # 평가손익_16이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_16 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_17 != None and irp_stock_nm_17 != '' :   # 상품명_17(퇴직 연금)이 존재하면
    rowNo = add_row + 24		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_17(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "17"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_14      # 01. 필드에 값 쓰기 ---> 상품명_14(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_14      # 01. 필드에 값 쓰기 ---> 평가금액_14(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_14) )

    if irp_eval_profit_loss_14 != ''  :   # 평가손익_14이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_14, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_18(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_14     # 01. 필드에 값 쓰기 ---> 평가손익_14(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_14     # 01. 필드에 값 쓰기 ---> 손익률(%)_14(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_14 == ''  :   # 평가손익_14이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_14 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_18 != None and irp_stock_nm_18 != '' :   # 상품명_18(퇴직 연금)이 존재하면
    rowNo = add_row + 25		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_18(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "18"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_18      # 01. 필드에 값 쓰기 ---> 상품명_18(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_18      # 01. 필드에 값 쓰기 ---> 평가금액_18(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_18) )

    if irp_eval_profit_loss_18 != ''  :   # 평가손익_18이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_18, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_18(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_18     # 01. 필드에 값 쓰기 ---> 평가손익_18(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_18     # 01. 필드에 값 쓰기 ---> 손익률(%)_18(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_18 == ''  :   # 평가손익_18이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_18 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_19 != None and irp_stock_nm_19 != '' :   # 상품명_19(퇴직 연금)이 존재하면
    rowNo = add_row + 26		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_19(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "19"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_19      # 01. 필드에 값 쓰기 ---> 상품명_19(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_19      # 01. 필드에 값 쓰기 ---> 평가금액_19(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_19) )

    if irp_eval_profit_loss_19 != ''  :   # 평가손익_19이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_19, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_19(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_19     # 01. 필드에 값 쓰기 ---> 평가손익_19(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_19     # 01. 필드에 값 쓰기 ---> 손익률(%)_19(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_19 == ''  :   # 평가손익_19이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_19 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->

if irp_stock_nm_20 != None and irp_stock_nm_20 != '' :   # 상품명_20(퇴직 연금)이 존재하면
    rowNo = add_row + 27		# "03-1. 퇴직신탁 보유현황" 칼럼 Row(02. 총자산(24Y) Tab에서)   ---> 상품명_20(퇴직 연금)
    wsTot[f'K{rowNo}'].value = "20"  # 01. 필드에 값 쓰기 ---> No. ■ 
    wsTot[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    wsTot[f'L{rowNo}'].value = "010-30-7208980"   # 01. 필드에 값 쓰기 ---> 계좌번호 ■ 
    wsTot[f'L{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'L{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'L{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'M{rowNo}'].value = "개인형IRP"   # 01. 필드에 값 쓰기 ---> 제도 ■ 
    wsTot[f'M{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'M{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'M{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'N{rowNo}'].value = irp_stock_nm_20      # 01. 필드에 값 쓰기 ---> 상품명_20(퇴직 연금) Value ■
    wsTot[f'N{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
    wsTot[f'N{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 

    wsTot[f'O{rowNo}'].value = irp_eval_amt_20      # 01. 필드에 값 쓰기 ---> 평가금액_20(퇴직 연금) Value ■
    wsTot[f'O{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'O{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'O{rowNo}'].border = thin_border   # 04. 필드 테두리 설정 
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_22] [평가손익_2(퇴직 연금)]"+ str(irp_eval_profit_loss_20) )

    if irp_eval_profit_loss_20 != ''  :   # 평가손익_20이 존재하면 
        wsTot[f'P{rowNo}'].value = format(irp_eval_profit_loss_20, ',')      # 01. 필드에 값 쓰기 ---> 평가손익_20(퇴직 연금) Value ■
    else: 
        wsTot[f'P{rowNo}'].value = irp_eval_profit_loss_20     # 01. 필드에 값 쓰기 ---> 평가손익_20(퇴직 연금) Value ■

    wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'P{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정 
    wsTot[f'P{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_51_23] [손익률(%)_2(퇴직 연금)]"+ str(irp_eval_profit_loss_rate_1) )

    wsTot[f'Q{rowNo}'].value = irp_eval_profit_loss_rate_20     # 01. 필드에 값 쓰기 ---> 손익률(%)_20(퇴직 연금) Value ■
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정  
    wsTot[f'Q{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')     # 03. 필드 정렬 설정
    wsTot[f'Q{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

    if irp_eval_profit_loss_20 == ''  :   # 평가손익_20이 존재하면     
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9)   # 02. 필드 글자 설정 
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정   
    elif irp_eval_profit_loss_20 > 0  :   # 평평가손익_2이 양이면    
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
        wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    else :
        wsTot[f'P{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    wsTot[f'Q{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
# # ------------------------------------------------------------------------------------------------------------->
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


print("\n\n [@_T] ■■■ [/ast_vrfc.py] ==> [T_01] ■■■■■■ [######################### [71. 자산(24Y) Tab 작업 Start] #########################] ■■■■■■ ")

print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_01] [71. 자산(24Y) Tab Sheet(Tab)에 접근 @@@@@@@@ ===========>]")

tot_sum_prev_amt = 0        # 90. 총 합계 금액(이전 년월)
rowNo = 17
astYear = "01. 자산("+ str(astYYM[2:4]) +"Y)"   # 02. 출력할 년월■  --> 자산(25Y)

if int(input_astYYYYMM[1]) > 11  :   # 자산 월이 12월 이면
    ws = wb.create_sheet(astYear, 0)   # 엑셀 Sheet 생성 --> 자산(25Y) Tab 생성
    ws = wb[astYear]    # "자산(2023)" Sheet(Tab)에 접근 --> 자산(25Y) Tab 접근    
    astPrevYear = "01. 자산("+ str(input_astYYYYMM[0][2:4]) +"Y)"   # 01. 현재 년월[입력] --> 자산(24Y)
    # print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_01_1] [00. astYear]"+ str(astYear) +"[자산 년]"+ str(input_astYYYYMM[0]) +"[자산 월]"+ str(input_astYYYYMM[1]) ) 
    wsPrev = wb[astPrevYear]    # 01. 현재 년월[입력] Sheet(Tab)에 접근 --> 자산(24Y) Tab 접근   

    tot_sum_prev_amt = wsPrev[f'B{rowNo}'].value    # 01. 현재 년월(Tab) 총 합계 금액(이전 년월) value ■  
else :    
    ws = wb[astYear]    # "자산(2023)" Sheet(Tab)에 접근 --> 자산(25Y) Tab 접근    
    tot_sum_prev_amt = ws[f'B{rowNo}'].value    # 01. 필드에 값 쓰기 ---> 00. 총 합계 금액(이전 년월) value ■ 
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_03] [00. astYear]"+ str(astYear) +"[자산 년]"+ str(input_astYYYYMM[0]) +"[자산 월]"+ str(input_astYYYYMM[1]) )
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_04] [rowNo]"+ str(rowNo) +"[00. 총 합계 금액(이전 년월)]"+ str(tot_sum_prev_amt) )
# -------------------------------------------------------------------------------------------->


# 01. 자산(24Y) Tab 줄 추가, 셀 병합
insertRow = 17 + 2      # 월별 자산 추가 줄(17줄 + 2줄[상위 빈 줄])
ws.insert_rows(3, insertRow)    # 월별 자산 제목 아래 부터 줄부터 18줄 추가(3번째 줄 위치에 19줄을 추가 ==> 17줄)
print("\n\n[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_04] [01.자산(24Y) Tab 줄 추가, 셀 병합 처리]===========>" )

addRow = 3
addRow2 = insertRow
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)  # 'TAMA 자산(2024)' 셀 병합("A1:E1")
ws.merge_cells(start_row=addRow, start_column=1, end_row=addRow, end_column=4)     # '2024.06 자산' 셀 병합("A3:D3") ==> 22
ws.merge_cells(start_row=addRow, start_column=7, end_row=addRow, end_column=8)     # '은행 결산' 셀 병합(G3:H3")
ws.merge_cells(start_row=addRow, start_column=10, end_row=addRow, end_column=11)   # '지출' 셀 병합("J3:K3")
ws.merge_cells(start_row=addRow2, start_column=1, end_row=addRow2, end_column=5)   # '결산' 셀 병합("A19:E19")
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_50] [item_0번째] [00. A_rowNo]"+ str(ws[f'A{addRow}'].value) +"[01. start_row]"+ str(addRow) +"[02. addRow2]"+ str(addRow2) )

addRow = 0
addRow2 = 0
addRowFrst = int(insertRow) + 3    # 줄 추가_01[2024.06 자산 칼럼] = 19 + 3줄[하위 빈 줄])
addRowFrst2 = int(insertRow) + 19  # 줄 추가_02[결산 칼럼] = 19 + 19줄[하위 빈 줄])

for i in range(1, 13):
    addRow = int(addRow) + int(addRowFrst)     # 줄 추가_03 = 22 + 3 = 41
    addRow2 = int(addRow2) + int(addRowFrst2)  # 줄 추가_03 = 19 + 19 = 57
    if i > 1 : addRow = int(addRow) - 3
    if i > 1 : addRow2 = int(addRow2) - 19
    # if ws[f'A{rowNo}'].value != None and ws[f'A{rowNo}'].value != '' :   # 상품명_3(퇴직 연금)이 존재하면 
    ws.merge_cells(start_row=addRow, start_column=1, end_row=addRow, end_column=4)     # '2024.06 자산' 셀 병합("A3:D3") ==> 41
    ws.merge_cells(start_row=addRow, start_column=7, end_row=addRow, end_column=8)     # '은행 결산' 셀 병합(G3:H3")
    ws.merge_cells(start_row=addRow, start_column=10, end_row=addRow, end_column=11)   # '지출' 셀 병합("J3:K3") 
    ws.merge_cells(start_row=addRow2, start_column=1, end_row=addRow2, end_column=5)   # '결산' 셀 병합("A19:E19")
    print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_51] [item_번째]"+ str(i) +"[00. A_rowNo]"+ str(ws[f'A{rowNo}'].value) +"[01. start_row]"+ str(addRow) +"[02. addRow2]"+ str(addRow2) )
# -------------------------------------------------------------------------------------------->

# 01. 자산(24Y) Tab 값 쓰기
ws.column_dimensions["A"].width = 50  # A열의 너비를 50로 설정
ws.column_dimensions["B"].width = 10  # A열의 너비를 10로 설정
ws.column_dimensions["G"].width = 18  # G열의 너비를 18로 설정
ws.column_dimensions["J"].width = 18  # G열의 너비를 18로 설정 
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

rowNo = 0
add_row = 0     # 줄 추가
rowNo = add_row + 1 
if int(input_astYYYYMM[1]) > 11  :   # 자산 월이 12월 이면
    ws[f'A{rowNo}'].value = "01. TAMA 자산("+ str(astYYM[0:4]) +")"  # 01. 필드에 값 쓰기 ---> TAMA 자산 ■
    ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=13)   # 02. 필드 글자 설정 
    ws[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
    ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
    ws[f'A{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 3     # "2023.12 자산" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = astYYM +" 자산"   # 01. 필드에 값 쓰기 ---> 자산 년월 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=12)   # 02. 필드 글자 설정 
ws[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = orangeFill      # 05. 필드 배경색 설정

ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = orangeFill      # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = orangeFill      # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = orangeFill      # 05. 필드 배경색 설정

ws[f'E{rowNo}'].value = lastLastDt     # 01. 필드에 값 쓰기 ---> 024.01.01(작성 년월일) ■
ws[f'E{rowNo}'].font = Font(name="돋움", bold=True, size=9)     # 02. 필드 글자 설정 
ws[f'E{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')    # 03. 필드 정렬 설정
ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = orangeFill      # 05. 필드 배경색 설정

ws[f'G{rowNo}'].value = '은행 결산'     # 01. 필드에 값 쓰기 ---> 은행 결산 ■
ws[f'G{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")     # 02. 필드 글자 설정
ws[f'G{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')   # 03. 필드 정렬 설정
ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'G{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'J{rowNo}'].value = '지출'          # 01. 필드에 값 쓰기 ---> 지출 ■
ws[f'J{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")      # 02. 필드 글자 설정
ws[f'J{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')    # 03. 필드 정렬 설정
ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'J{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'K{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정 
# ------------------------------------------------------------------------------------------------------------->

rowNo = add_row + 4    # "내용" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '내용'          # 01. 필드에 값 쓰기 ---> 내용 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'A{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'B{rowNo}'].value = '금액'          # 01. 필드에 값 쓰기 ---> 금액 ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'C{rowNo}'].value = '원금'          # 01. 필드에 값 쓰기 ---> 원금 ■
ws[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'C{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'D{rowNo}'].value = '손익'          # 01. 필드에 값 쓰기 ---> 손익 ■
ws[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'D{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'E{rowNo}'].value = '수익률(%)'     # 01. 필드에 값 쓰기 ---> 수익률(%) ■
ws[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'E{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'G{rowNo}'].value = '은행 자산 상태'    # 01. 필드에 값 쓰기 ---> 은행 자산 상태 ■
ws[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'G{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'G{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'H{rowNo}'].value = '금액'    # 01. 필드에 값 쓰기 ---> 금액 ■
ws[f'H{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'H{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'H{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'J{rowNo}'].value = '내역'    # 01. 필드에 값 쓰기 ---> 내역 ■
ws[f'J{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'J{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'J{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정

ws[f'K{rowNo}'].value = '금액'    # 01. 필드에 값 쓰기 ---> 금액 ■
ws[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'K{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')     # 03. 필드 정렬 설정
ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'K{rowNo}'].fill = grayFill        # 05. 필드 배경색 설정
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 5    # "01. 은행 예. 적금([06])" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = deposit_savings_nm    # 01. 필드에 값 쓰기 ---> 01. 은행 예. 적금([06]) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'B{rowNo}'].value = format(deposit_savings, ',')         # 01. 필드에 값 쓰기 ---> 01. 은행 예. 적금([06]) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'G{rowNo}'].value = deposit_savings_sh_nm   # 01. 필드에 값 쓰기 ---> 01. 신한은행 ■
ws[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'H{rowNo}'].value = format(deposit_savings_sh, ',')     # 01. 필드에 값 쓰기 ---> 01. 신한은행 Value ■
ws[f'H{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정 
ws[f'H{rowNo}'].border = thin_border     # 04. 필드 테두리 설정
ws[f'H{rowNo}'].fill = blueFill          # 05. 필드 배경색 설정

ws[f'J{rowNo}'].value = '91. 월 지출(A)'   # 01. 필드에 값 쓰기 --->  91. 월 지출(A)■
ws[f'J{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'K{rowNo}'].value = format(mon_expendit, ',')     # 01. 필드에 값 쓰기 --->  91. 월 지출(A) Value ■
ws[f'K{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'K{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정 설정
ws[f'K{rowNo}'].border = thin_border     # 04. 필드 테두리 설정
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 6    # "02. 대여금(문태용[24.05까지])" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = rental_fee_nm   # 01. 필드에 값 쓰기 ---> 02. 대여금(문태용[24.05까지]) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_11] [02. 대여금]"+ str(rental_fee))

ws[f'B{rowNo}'].value = format(rental_fee, ',')        # 01. 필드에 값 쓰기 ---> 02. 대여금(문태용[24.05까지]) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = blueFill        # 05. 필드 배경색 설정

ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# # -------------------------------------------------------------------------------------------->

rowNo = add_row + 7   # "1. 총 예적금" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '1. 총 예적금'   # 01. 필드에 값 쓰기 ---> 1. 총 예적금 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

ws[f'B{rowNo}'].value = format(tot_deposit_savings, ',')         # 01. 필드에 값 쓰기 ---> 01. 은행 예. 적금([06]) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = greenWeekFill   # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 8   # "77. 주식 투자" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '77. 주식 투자'   # 01. 필드에 값 쓰기 ---> 77. 주식 투자 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'B{rowNo}'].value = format(eval_amt, ',')         # 01. 필드에 값 쓰기 ---> 77. 주식 투자(평가금액) value  ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = yellowFill      # 05. 필드 배경색 설정

ws[f'C{rowNo}'].value = format(prchs_amt, ',')         # 01. 필드에 값 쓰기 ---> 77. 주식 투자 원금(매입금액) value ■
ws[f'C{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
ws[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

# 평가손익(원) = 평가금액 - 매입금액
eval_profit_loss =  eval_amt - prchs_amt
ws[f'D{rowNo}'].value = format(eval_profit_loss, ',')         # 01. 필드에 값 쓰기 ---> 77. 주식 투자 손익(평가손익) value ■
if eval_profit_loss > 0  :   # 77. 주식 투자 손익이 양이면
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
ws[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

# 손익률(%)= (평가금액 - 매입금액) / 매입금액
eval_profit_loss_rate =  (eval_profit_loss / prchs_amt) * 100
ws[f'E{rowNo}'].value = f'{eval_profit_loss_rate:.2f}'         # 01. 필드에 값 쓰기 ---> 77. 주식 투자 수익률(%) [손익률(%)] value ■ 
ws[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 9   # "90. 퇴직 연금(개인형 IRP)'"칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '90. 퇴직 연금(개인형 IRP)'   # 01. 필드에 값 쓰기 ---> 90. 퇴직 연금(개인형 IRP) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'B{rowNo}'].value = format(irp_eval_amt, ',')         # 01. 필드에 값 쓰기 ---> 90. 퇴직 연금(개인형 IRP) (평가금액) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = blueDarkFill    # 05. 필드 배경색 설정

ws[f'C{rowNo}'].value = format(irp_prchs_amt, ',')         # 01. 필드에 값 쓰기 ---> 78. 퇴직 연금 원금(매입금액) value ■
ws[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'D{rowNo}'].value = format(irp_eval_profit_loss, ',')         # 01. 필드에 값 쓰기 ---> 78. 퇴직 연금 손익(평가손익) value ■
if irp_eval_profit_loss > 0  :   # 78. 퇴직 연금 손익이 양이면
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
ws[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_13] [퇴직 연금 수익률(%)]"+ str(irp_eval_profit_loss_rate))

ws[f'E{rowNo}'].value = format(irp_eval_profit_loss_rate * 100, ',')    # 01. 필드에 값 쓰기 ---> 78. 퇴직 연금 수익률(%) value ■
ws[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 10   # "78. 발행어음 만기형_개인(181~270)'" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '78. 발행어음 만기형_개인(181~270)'   # 01. 필드에 값 쓰기 ---> 78. 발행어음 만기형_개인(181~270) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'B{rowNo}'].value = format(eval_amt78, ',')         # 01. 필드에 값 쓰기 ---> 78. 발행어음 만기형_개인(181~270)(평가금액) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = blueDarkFill    # 05. 필드 배경색 설정

ws[f'C{rowNo}'].value = format(prchs_amt78, ',')         # 01. 필드에 값 쓰기 ---> 78. 발행어음 원금(매입금액) value ■
ws[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'C{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

# 평가손익(원) = 평가금액 - 매입금액
eval_profit_loss78 = eval_amt78 - prchs_amt78
ws[f'D{rowNo}'].value = format(eval_profit_loss78, ',')   # 01. 필드에 값 쓰기 ---> 78. 발행어음 손익(평가손익) value ■
if eval_profit_loss78 > 0  :   # 8. 발행어음 손익이 양이면
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색) 
ws[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

# 손익률(%)= (평가금액 - 매입금액) / 매입금액
eval_profit_loss_rate78 = (eval_profit_loss78 / prchs_amt78) * 100
# print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_12] [발행어음 수익률(%)]"+ str(eval_profit_loss_rate))

ws[f'E{rowNo}'].value = f'{eval_profit_loss_rate78:.2f}'      # 01. 필드에 값 쓰기 ---> 78. 발행어음 수익률(%) value ■ 
ws[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 11   # "79. 발행어음 CMA_개인'" 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '79. 발행어음 CMA_개인'   # 01. 필드에 값 쓰기 ---> 79. 발행어음 CMA_개인 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = blueDarkFill    # 05. 필드 배경색 설정
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_11] [cma_prchs_amt79]"+ str(ws_03_bond[f'E{rowNo}'].value))

ws[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정

# 평가손익(원) = 평가금액 - 매입금액
eval_profit_loss79 = cma_eval_amt79 - cma_prchs_amt79
if eval_profit_loss79 > 0  :   # 8. 발행어음 손익이 양이면
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정(빨간색)
else :
    ws[f'D{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색)
    ws[f'E{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="00ADEF")   # 02. 필드 글자 설정(파란색) 
ws[f'D{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

# 손익률(%)= (평가금액 - 매입금액) / 매입금액 
if eval_profit_loss79 < 1 :   # 발행어음 수익률이 음이면
    eval_profit_loss_rate79 = 0
else :
    eval_profit_loss_rate79 =  (eval_profit_loss79 / cma_prchs_amt79) * 100
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_12] [발행어음 수익률(%)]"+ str(eval_profit_loss_rate))
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_13] [05. 평가금액(79. 발행어음 CMA_개인)]"+ str(ws_03_bond[f'E{cmaRowNo}'].value))

if ws_03_bond[f'E{cmaRowNo}'].value != None  :  # 05. 평가금액(79. 발행어음 CMA_개인)이 널이 아니면
    ws[f'B{rowNo}'].value = format(cma_eval_amt79, ',')    # 79. 발행어음 CMA_개인(01. 금액) value ■
    ws[f'C{rowNo}'].value = format(cma_prchs_amt79, ',')    # 79. 발행어음 CMA_개인(02. 원금) value ■
    ws[f'D{rowNo}'].value = format(eval_profit_loss79, ',')   # 79. 발행어음 CMA_개인(03. 손익) value ■
    ws[f'E{rowNo}'].value = f'{eval_profit_loss_rate79:.2f}'  # 79. 발행어음 CMA_개인(03. 수익률(%)) value ■
else:
    ws[f'B{rowNo}'].value = ""   # 79. 발행어음 CMA_개인(01. 금액) value ■
    ws[f'C{rowNo}'].value = ""   # 79. 발행어음 CMA_개인(02. 원금) value ■
    ws[f'D{rowNo}'].value = ""   # 79. 발행어음 CMA_개인(03. 손익) value ■
    ws[f'E{rowNo}'].value = ""  # 79. 발행어음 CMA_개인(03. 수익률(%)) value ■
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_14] [05. 평가금액(79. 발행어음 CMA_개인)]"+ str(ws[f'B{rowNo}'].value))

ws[f'E{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'E{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 12   # "2. 총 투자 금액" 칼 럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '2. 총 투자 금액'   # 01. 필드에 값 쓰기 ---> 2. 총 투자 금액 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

tot_invst_amt = int(eval_amt) + int(eval_amt78) + int(cma_eval_amt79) + int(irp_eval_amt)   # 2. 총 투자 금액 = 77. 평가금액(주식 투자) + 78. 평가금액(발행 어음) + 79. 발행어음 CMA_개인 + 90. 평가금액(퇴직 연금)
ws[f'B{rowNo}'].value = format(tot_invst_amt, ',')         # 01. 필드에 값 쓰기 ---> 2. 총 투자 금액 value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

tot_prch_amt = int(prchs_amt) + int(prchs_amt78) + int(cma_prchs_amt79) + int(irp_prchs_amt)   # 2. 총 투자 금액 (원금)= 77. 평가금액(주식 투자)(원금) + 78. 평가금액(발행 어음) (원금)+ 90. 평가금액(퇴직 연금)(원금)
ws[f'C{rowNo}'].value = format(tot_prch_amt, ',')         # 01. 필드에 값 쓰기 ---> 2. 총 투자 금액(원금) value ■
ws[f'C{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'C{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = greenWeekFill   # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = greenWeekFill    # 05. 필드 배경색 설정

ws[f'G{rowNo}'].border = thin_border    # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 13   # "21. (무) 굿앤굿어린이CI보험(Hi1304)(현대해상, 진수종) 칼럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '21. (무) 굿앤굿어린이CI보험(Hi1304)(현대해상, 진수종)'   # 01. 필드에 값 쓰기 ---> 21. (무) 굿앤굿어린이CI보험(Hi1304)(현대해상, 진수종) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'B{rowNo}'].value = format(insu_non_life_soo, ',')         # 01. 필드에 값 쓰기 ---> 21. (무) 굿앤굿어린이CI보험(Hi1304)(현대해상, 진수종) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정

ws[f'G{rowNo}'].border = thin_border    # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# # -------------------------------------------------------------------------------------------->

rowNo = add_row + 14   # "22. (무) 한아름행복플러스종합보험1404(한화손해보험, 잔태만)'" 칼 럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '22. (무) 한아름행복플러스종합보험1404(한화손해보험, 잔태만)'   # 01. 필드에 값 쓰기 ---> 2. (무) 한아름행복플러스종합보험1404(한화손해보험, 잔태만) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'B{rowNo}'].value = format(insu_non_life, ',')         # 01. 필드에 값 쓰기 ---> 22. (무) 한아름행복플러스종합보험1404(한화손해보험, 잔태만) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=False, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정

ws[f'G{rowNo}'].border = thin_border    # 04. 필드 테두리 설정  ---> 은행 결산_1 칼럼 ■

ws[f'H{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 은행 결산_2 칼럼 ■

ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_1 칼럼 ■

ws[f'K{rowNo}'].border = thin_border   # 04. 필드 테두리 설정  ---> 지출_2 칼럼 ■
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 15   # "2. 총 보험금" 칼 럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '2. 총 보험금'   # 01. 필드에 값 쓰기 ---> 2. 총 보험금 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = greenFill    # 05. 필드 배경색 설정

tot_insu_non_life = int(insu_non_life) + int(insu_non_life)   # 2. 총 보험금 -------->
ws[f'B{rowNo}'].value = format(tot_insu_non_life, ',')         # 01. 필드에 값 쓰기 ---> 2. 총 보험금 value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = greenFill    # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = greenFill    # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = greenFill   # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = greenFill    # 05. 필드 배경색 설정

ws[f'G{rowNo}'].value = '[06] 은행 예. 적금 총합'   # 01. 필드에 값 쓰기 ---> [06] 은행 예. 적금 총합 ■
ws[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9, color="FF0000")   # 02. 필드 글자 설정
ws[f'G{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'G{rowNo}'].fill = blueDark2Fill   # 05. 필드 배경색 설정

ws[f'H{rowNo}'].value = format(deposit_savings, ',')      # 01. 필드에 값 쓰기 ---> [06] 은행 예. 적금 총합 Value ■
ws[f'H{rowNo}'].font =Font(name="돋움", bold=False, size=9, color="FF0000")    # 02. 필드 글자 설정
ws[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정 
ws[f'H{rowNo}'].border = thin_border     # 04. 필드 테두리 설정
ws[f'H{rowNo}'].fill = blueDark2Fill     # 05. 필드 배경색 설정

ws[f'J{rowNo}'].value = '지출(A) 총합'   # 01. 필드에 값 쓰기 ---> 지출(A) 총합 ■
ws[f'J{rowNo}'].font = Font(name="돋움", bold=True, size=9)   # 02. 필드 글자 설정
ws[f'J{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'J{rowNo}'].fill = greenFill       # 05. 필드 배경색 설정

ws[f'K{rowNo}'].value = format(mon_expendit, ',')      # 01. 필드에 값 쓰기 ---> 지출(A) 총합 Value ■
ws[f'K{rowNo}'].font =Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정
ws[f'K{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정 
ws[f'K{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'K{rowNo}'].fill = greenFill        # 05. 필드 배경색 설정 
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 16   # "3. 부동산(현아트빌 404호)" 칼 럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '3. 부동산(현아트빌 404호)'   # 01. 필드에 값 쓰기 ---> 3. 부동산(현아트빌 404호) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = greenFill        # 05. 필드 배경색 설정

ws[f'B{rowNo}'].value = format(self_house, ',')         # 01. 필드에 값 쓰기 ---> 3. 부동산(현아트빌 404호) value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=True, size=9)    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = greenFill        # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = greenFill        # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = greenFill        # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = greenFill        # 05. 필드 배경색 설정 
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 17   # "90. 총 합계" 칼 럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '90. 총 합계'   # 01. 필드에 값 쓰기 ---> 90. 총 합계 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

tot_sum_amt = int(tot_deposit_savings) + int(tot_invst_amt) + int(tot_insu_non_life) + int(self_house)   
# 90. 총 합계 금액 = 1. 총 예적금 + 2. 총 투자 금액 + 3. 총 보험금 + 4. 부동산(현아트빌 404호)
ws[f'B{rowNo}'].value = format(tot_sum_amt, ',')         # 01. 필드에 값 쓰기 ---> 90. 총 합계 value ■
ws[f'B{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = yellowFill   # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

ws[f'G{rowNo}'].value = '총합계 전월 대비 증감'   # 01. 필드에 값 쓰기 ---> 총합계 전월 대비 증감 ■
ws[f'G{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'G{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')    # 03. 필드 정렬 설정
ws[f'G{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'G{rowNo}'].fill = grayFill    # 05. 필드 배경색 설정

ws[f'H{rowNo}'].value =  '금액'          # 01. 필드에 값 쓰기 ---> 금액 ■
ws[f'H{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'H{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')    # 03. 필드 정렬 설정
ws[f'H{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'H{rowNo}'].fill = grayFill    # 05. 필드 배경색 설정

ws[f'J{rowNo}'].value = '수입(B) 총합'   # 01. 필드에 값 쓰기 ---> 수입(B) 총합 ■
ws[f'J{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정 
ws[f'J{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'J{rowNo}'].fill = grayDarkFill    # 05. 필드 배경색 설정

ws[f'K{rowNo}'].value = format(mon_income, ',')         # 01. 필드에 값 쓰기 ---> 수입(B) 총합 value ■
ws[f'K{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")   # 02. 필드 글자 설정
ws[f'K{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'K{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'K{rowNo}'].fill = grayDarkFill    # 05. 필드 배경색 설정
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 18   # "91. 가용 자산" 칼 럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = '91. 가용 자산'   # 01. 필드에 값 쓰기 ---> 91. 가용 자산 ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

# ws.unmerge_cells("A18:D18")     # 셀 병합 해제 

available_assets = int(tot_deposit_savings) + int(tot_invst_amt)    # 1. 가용 자산 = 1. 총 예적금 + 2. 총 투자 금액 
print("\n\n[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_11_2] [91. 가용 자산_Row]"+ str(rowNo) +"[91. 가용 자산]"+ str(available_assets) )

ws[f'B{rowNo}'].value = format(available_assets, ',')    # 01. 필드에 값 쓰기 ---> 91. 가용 자산 value ■  
ws[f'B{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정
ws[f'B{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'B{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = yellowFill   # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

ws[f'G{rowNo}'].value = '총 합계(90.A - 90.B)'   # 01. 필드에 값 쓰기 ---> 총 합계 ■
ws[f'G{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'G{rowNo}'].alignment = Alignment(horizontal='center', vertical='center')    # 03. 필드 정렬 설정
ws[f'G{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'G{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정 

if tot_sum_prev_amt == None or tot_sum_prev_amt == '' :   # 90. 총 합계 금액(이전 년월)이 널 이면 ■■■
    tot_sum_prev_amt = 0
else:
    tot_sum_prev_amt = int(tot_sum_prev_amt.replace(',',''))     # 총 합계 금액(콤마제거 후 정수로 처리)     
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_12] [00. 총 합계 금액]"+ str(tot_sum_prev_amt) )

rowNo_prev = add_row + 16
tot_sum_for_mon_amt = int(tot_sum_amt) - int(tot_sum_prev_amt)  # 전월 대비 증감 총 합계 = 90. 총 합계 금액(당월) -  90. 총 합계 금액(이전 년월)
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_71_15] [90. 총 합계(90.A - 90.B)]"+ str(tot_sum_for_mon_amt) +"[90. 총 합계 금액(당월)]"+ str(tot_sum_amt) +"[90. 총 합계 금액(이전 년월)]"+ str(tot_sum_prev_amt) )

ws[f'H{rowNo}'].value = format(tot_sum_for_mon_amt, ',')         # 01. 필드에 값 쓰기 ---> 총 합계(90.A - 90.B) value ■
ws[f'H{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
ws[f'H{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'H{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'H{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

ws[f'J{rowNo}'].value = '총 합계([B] - [A])'   # 01. 필드에 값 쓰기 ---> 총 합계([B] - [A])■
ws[f'J{rowNo}'].font = Font(name="돋움", bold=True, size=9)  # 02. 필드 글자 설정
ws[f'J{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'J{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정

totSum = int(mon_income) - int(mon_expendit)
ws[f'K{rowNo}'].value = format(totSum, ',')         # 01. 필드에 값 쓰기 ---> 총 합계([B] - [A]) value ■
ws[f'K{rowNo}'].font = Font(name="돋움", bold=True, size=9, color="FF0000")    # 02. 필드 글자 설정(빨간색)
ws[f'K{rowNo}'].alignment = Alignment(horizontal='right', vertical='center')    # 03. 필드 정렬 설정
ws[f'K{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'K{rowNo}'].fill = yellowFill    # 05. 필드 배경색 설정
# -------------------------------------------------------------------------------------------->

rowNo = add_row + 19   # "결산" 칼 럼 Row(자산(2023) Tab에서)
ws[f'A{rowNo}'].value = settlement   # 01. 필드에 값 쓰기 ---> 결산(코멘트) ■
ws[f'A{rowNo}'].font = Font(name="돋움", bold=False, size=9)  # 02. 필드 글자 설정
ws[f'A{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'A{rowNo}'].fill = orangeWeekFill    # 05. 필드 배경색 설정

ws[f'B{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'B{rowNo}'].fill = orangeWeekFill    # 05. 필드 배경색 설정

ws[f'C{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'C{rowNo}'].fill = orangeWeekFill    # 05. 필드 배경색 설정

ws[f'D{rowNo}'].border = thin_border   # 04. 필드 테두리 설정
ws[f'D{rowNo}'].fill = orangeWeekFill   # 05. 필드 배경색 설정

ws[f'E{rowNo}'].border = thin_border    # 04. 필드 테두리 설정
ws[f'E{rowNo}'].fill = orangeWeekFill    # 05. 필드 배경색 설정
# ===============================================================================================================================
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->


# 틀 고정
# ws.freeze_panes = "B2" # B2 기준으로 틀 고정

rsltFileNm = "02. 자산 검증("+ astYM +")_RSLT.xlsx"    # 결과 파일 명(02. 자산 검증(23.08)_RSLT.xlsx) 
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_91] [URL 경로]"+ str(urlPath) +"[결과 파일 명 ■■■■■■ ]"+ rsltFileNm  )

wb.save(urlPath + rsltFileNm)
# wb.save("02. 자산 검증(23.08)_RSLT.xlsx")
print("[@_T] ■■■ [/ast_vrfc.py] ==> [T_99] ■■■■■■ [######################### [자산 검증 파일 TEST End] #########################] ■■■■■■\n\n\n\n")
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■ ------------------->